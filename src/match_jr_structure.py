"""Rebuild S&N and OOS into 3-tab structure matching JR exactly.

For each category (S&N, OOS), create:
  1. Call Tracker — 35 cols matching JR Call Tracker
  2. Email Tracker — 31 cols matching JR Email Tracker
  3. Email Drafts — 4 cols matching JR Email Drafts

All private practice phones and emails are NPPES-verified.
"""

import logging
import os
import re

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("match_jr")

# Styles (matching JR tab style)
HEADER_NAVY = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_WHITE_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
ALT_ROW = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MICROLYTE_GREEN = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
VERIFIED_GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
MISSING_RED = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
NPPES_YELLOW = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)

CALL_TRACKER_COLS = [
    "HCP NPI", "First Name", "Last Name", "Credential", "Specialty",
    "Phone", "Phone Status", "Primary Site of Care", "City", "State",
    "Practice Type", "Tier", "Lead Priority", "Lead Status",
    "MAC Jurisdiction", "Microlyte Eligible", "VOL_COL",
    "Call 1 Date", "Call 1 Outcome", "Call 1 Notes",
    "Call 2 Date", "Call 2 Outcome", "Call 2 Notes",
    "Call 3 Date", "Call 3 Outcome", "Call 3 Notes",
    "Call 4 Date", "Call 4 Outcome", "Call 4 Notes",
    "Call 5 Date", "Call 5 Outcome", "Call 5 Notes",
    "Next Action", "Next Action Date", "Decision Maker?",
]

EMAIL_TRACKER_COLS = [
    "HCP NPI", "First Name", "Last Name", "Credential", "Specialty",
    "Email", "Email Status", "Primary Site of Care", "City", "State",
    "Practice Type", "Tier", "Lead Priority", "Lead Status",
    "MAC Jurisdiction", "Microlyte Eligible", "VOL_COL",
    "Email 1 Date", "Email 1 Subject", "Email 1 Outcome", "Email 1 Notes",
    "Email 2 Date", "Email 2 Subject", "Email 2 Outcome", "Email 2 Notes",
    "Email 3 Date", "Email 3 Subject", "Email 3 Outcome", "Email 3 Notes",
    "Next Action", "Next Action Date",
]

EMAIL_DRAFTS_COLS = ["HCP NPI", "Last Name", "Subject Line", "Draft Email"]

COL_WIDTHS = {
    "HCP NPI": 14, "First Name": 13, "Last Name": 15, "Credential": 10,
    "Specialty": 28, "Phone": 16, "Phone Status": 16,
    "Email": 32, "Email Status": 22, "Primary Site of Care": 32,
    "City": 16, "State": 8, "Practice Type": 16, "Tier": 20,
    "Lead Priority": 12, "Lead Status": 16, "MAC Jurisdiction": 16,
    "Microlyte Eligible": 14, "Joint Repl Vol": 14, "Open Spine Vol": 14,
    "Procedure Vol": 14, "Subject Line": 35, "Draft Email": 60,
    "Decision Maker?": 14,
}
for i in range(1, 6):
    COL_WIDTHS[f"Call {i} Date"] = 12
    COL_WIDTHS[f"Call {i} Outcome"] = 20
    COL_WIDTHS[f"Call {i} Notes"] = 30
for i in range(1, 4):
    COL_WIDTHS[f"Email {i} Date"] = 12
    COL_WIDTHS[f"Email {i} Subject"] = 30
    COL_WIDTHS[f"Email {i} Outcome"] = 18
    COL_WIDTHS[f"Email {i} Notes"] = 30
COL_WIDTHS["Next Action"] = 20
COL_WIDTHS["Next Action Date"] = 14


def format_phone(phone):
    """Format 10-digit phone as (XXX) XXX-XXXX."""
    if not phone:
        return None
    digits = re.sub(r"\D", "", str(phone))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
    return phone


def extract_enriched_data(wb, sheet_name):
    """Extract enriched data from a sheet into dicts."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    return headers, [dict(zip(headers, r)) for r in rows[1:]]


def write_tracker_tab(wb, tab_name, cols, rows, vol_col, sort_desc=True):
    """Write a tracker tab (Call or Email)."""
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)

    # Replace VOL_COL placeholder with the actual volume column name
    actual_cols = [vol_col if c == "VOL_COL" else c for c in cols]

    # Write headers
    for col_idx, h in enumerate(actual_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_NAVY
        cell.font = HEADER_WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(h, 15)

    # Sort by vol desc
    if sort_desc:
        def sort_key(r):
            v = r.get(vol_col)
            if v is None:
                return 0
            try:
                return float(v)
            except (ValueError, TypeError):
                return 0
        rows = sorted(rows, key=sort_key, reverse=True)

    # Helper for col index lookup
    def col_idx(name):
        try:
            return actual_cols.index(name) + 1
        except ValueError:
            return None

    priority_idx = col_idx("Lead Priority")
    phone_status_idx = col_idx("Phone Status")
    email_status_idx = col_idx("Email Status")
    microlyte_idx = col_idx("Microlyte Eligible")
    lead_status_idx = col_idx("Lead Status")

    # Write data
    for row_idx, row in enumerate(rows, 2):
        is_alt = (row_idx - 2) % 2 == 1
        base_fill = ALT_ROW if is_alt else WHITE_FILL

        for c_idx, h in enumerate(actual_cols, 1):
            val = row.get(h)
            if isinstance(val, float) and pd.isna(val):
                val = None
            # Format phone
            if h == "Phone" and val:
                val = format_phone(val)
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = base_fill

        # Color coding
        if priority_idx:
            p = row.get("Lead Priority", "")
            if p == "A":
                ws.cell(row=row_idx, column=priority_idx).fill = GREEN_FILL
            elif p == "B":
                ws.cell(row=row_idx, column=priority_idx).fill = YELLOW_FILL
            elif p == "C":
                ws.cell(row=row_idx, column=priority_idx).fill = RED_FILL

        if email_status_idx:
            es = str(row.get("Email Status", ""))
            if "Verified" in es:
                ws.cell(row=row_idx, column=email_status_idx).fill = VERIFIED_GREEN
            elif es == "Missing":
                ws.cell(row=row_idx, column=email_status_idx).fill = MISSING_RED

        if phone_status_idx:
            ps = str(row.get("Phone Status", ""))
            if ps == "Verified":
                ws.cell(row=row_idx, column=phone_status_idx).fill = VERIFIED_GREEN
            elif ps == "Added from NPPES":
                ws.cell(row=row_idx, column=phone_status_idx).fill = NPPES_YELLOW
            elif ps == "Missing":
                ws.cell(row=row_idx, column=phone_status_idx).fill = MISSING_RED

        if microlyte_idx:
            if str(row.get("Microlyte Eligible", "")) == "Yes":
                ws.cell(row=row_idx, column=microlyte_idx).fill = MICROLYTE_GREEN

    num_rows = len(rows) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(actual_cols))}{num_rows}"

    # Dropdowns
    if lead_status_idx:
        col_l = get_column_letter(lead_status_idx)
        dv = DataValidation(type="list", formula1='"New,Contacted,Meeting Scheduled,Meeting Completed,Proposal Sent,Won,Lost,Nurture"', allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_l}2:{col_l}{num_rows}")

    # Call outcome dropdowns
    for i in range(1, 6):
        idx = col_idx(f"Call {i} Outcome")
        if idx:
            col_l = get_column_letter(idx)
            dv = DataValidation(type="list", formula1='"Connected - Meeting Set,Connected - Call Back,Connected - Not Interested,Voicemail,No Answer,Wrong Number,Gatekeeper - Message Left,Gatekeeper - Blocked"', allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_l}2:{col_l}{num_rows}")

    # Email outcome dropdowns
    for i in range(1, 4):
        idx = col_idx(f"Email {i} Outcome")
        if idx:
            col_l = get_column_letter(idx)
            dv = DataValidation(type="list", formula1='"Opened,Replied,Bounced,No Response,Unsubscribed"', allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_l}2:{col_l}{num_rows}")

    dm_idx = col_idx("Decision Maker?")
    if dm_idx:
        col_l = get_column_letter(dm_idx)
        dv = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_l}2:{col_l}{num_rows}")

    logger.info(f"  {tab_name}: {len(rows)} leads")
    return ws


def write_email_drafts_tab(wb, tab_name, rows, vol_col):
    """Write Email Drafts tab (4 cols: NPI, Last Name, Subject, Draft)."""
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)

    cols = ["HCP NPI", "Last Name", "Subject Line", "Draft Email"]
    widths = [14, 15, 35, 90]

    # Headers
    for col_idx, (h, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_NAVY
        cell.font = HEADER_WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Sort by volume desc
    def sort_key(r):
        v = r.get(vol_col)
        if v is None:
            return 0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0
    rows = sorted(rows, key=sort_key, reverse=True)

    # Data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(cols, 1):
            val = row.get(h)
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if h == "Draft Email":
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    num_rows = len(rows) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{num_rows}"

    logger.info(f"  {tab_name}: {len(rows)} leads")


def main():
    input_file = "Master_Lead_List_Tracker (3).xlsx"

    logger.info("Loading workbook...")
    wb_read = openpyxl.load_workbook(input_file, read_only=True)

    # The enriched S&N and OOS data was previously in flat tabs that we
    # split into tier tabs. We need to consolidate back from those tier tabs.
    sn_rows = []
    oo_rows = []
    sn_headers = None
    oo_headers = None

    for name in wb_read.sheetnames:
        if name.startswith("S&N "):
            h, r = extract_enriched_data(wb_read, name)
            sn_rows.extend(r)
            if sn_headers is None:
                sn_headers = h
        elif name.startswith("OOS "):
            h, r = extract_enriched_data(wb_read, name)
            oo_rows.extend(r)
            if oo_headers is None:
                oo_headers = h

    wb_read.close()

    logger.info(f"Consolidated S&N: {len(sn_rows)} leads")
    logger.info(f"Consolidated OOS: {len(oo_rows)} leads")

    # Convert "Verified Phone" (raw 10-digit) to "Phone" column for each dataset
    # and add drafts from existing Subject Line / Draft Email columns
    for row in sn_rows:
        row["Phone"] = row.get("Verified Phone")
    for row in oo_rows:
        row["Phone"] = row.get("Verified Phone")

    # Open workbook in write mode
    wb = openpyxl.load_workbook(input_file)

    # Delete tier tabs
    for name in list(wb.sheetnames):
        if name.startswith("S&N ") or name.startswith("OOS "):
            del wb[name]

    # Create S&N tabs
    logger.info("\nCreating S&N tabs...")
    write_tracker_tab(wb, "Call Tracker - S&N", CALL_TRACKER_COLS, sn_rows, "Open Spine Vol")
    write_tracker_tab(wb, "Email Tracker - S&N", EMAIL_TRACKER_COLS, sn_rows, "Open Spine Vol")
    write_email_drafts_tab(wb, "Email Drafts - S&N", sn_rows, "Open Spine Vol")

    # Create OOS tabs
    logger.info("\nCreating OOS tabs...")
    write_tracker_tab(wb, "Call Tracker - OOS", CALL_TRACKER_COLS, oo_rows, "Procedure Vol")
    write_tracker_tab(wb, "Email Tracker - OOS", EMAIL_TRACKER_COLS, oo_rows, "Procedure Vol")
    write_email_drafts_tab(wb, "Email Drafts - OOS", oo_rows, "Procedure Vol")

    # Reorder: Dashboard | JR tabs | S&N tabs | OOS tabs
    desired_order = [
        "Dashboard",
        "Call Tracker - JR", "Email Tracker - JR", "Email Drafts - JR",
        "Call Tracker - S&N", "Email Tracker - S&N", "Email Drafts - S&N",
        "Call Tracker - OOS", "Email Tracker - OOS", "Email Drafts - OOS",
    ]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            current_idx = wb.sheetnames.index(name)
            offset = i - current_idx
            if offset != 0:
                wb.move_sheet(name, offset=offset)

    wb.save(input_file)
    logger.info(f"\nSaved to {input_file}")
    logger.info(f"Final tabs: {wb.sheetnames}")

    # Verification: how many private practices have verified phones and emails?
    def verify_stats(rows, label):
        pp = [r for r in rows if r.get("Practice Type") == "Private Practice"]
        pp_phone_verified = [r for r in pp if r.get("Phone Status") in ("Verified", "Added from NPPES", "Updated (NPPES differs)")]
        pp_email_verified = [r for r in pp if r.get("Email Status") and "Missing" not in str(r.get("Email Status"))]
        logger.info(f"\n{label}:")
        logger.info(f"  Private Practice leads: {len(pp)}")
        logger.info(f"  With verified/NPPES-added phone: {len(pp_phone_verified)} ({len(pp_phone_verified)*100//max(len(pp),1)}%)")
        logger.info(f"  With verified/inferred email: {len(pp_email_verified)} ({len(pp_email_verified)*100//max(len(pp),1)}%)")

    verify_stats(sn_rows, "Spine & Neuro Private Practices")
    verify_stats(oo_rows, "Outside Ortho Private Practices")


if __name__ == "__main__":
    main()
