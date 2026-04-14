"""Add collagen volume and incision likelihood columns to all 3 Call Trackers + update Dashboard."""

import logging
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("add_collagen")

HEADER_NAVY = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_ROW = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

INCISION_FILLS = {
    "High": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),        # dark green
    "Medium-High": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),  # light green
    "Medium": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),        # amber
    "Low": PatternFill(start_color="F79646", end_color="F79646", fill_type="solid"),           # orange
    "Unlikely": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),      # red
}


def build_collagen_lookup_jr():
    """Build NPI → collagen data lookup for JR leads from source CSV."""
    df = pd.read_csv("data/input/collagen_dme_targeting.csv", dtype=str)
    df.columns = [c.strip('"') for c in df.columns]

    lookup = {}
    for _, row in df.iterrows():
        npi = str(row.get("HCP NPI", "")).strip()
        if not npi:
            continue
        lookup[npi] = {
            "Lg Collagen Vol": row.get("Lg Collagen Sheet - Procedure Volume"),
            "Sm/Md Collagen Vol": row.get("Sm/Md Collagen Sheet,Collagen Powder,Lg Collagen Sheet - Procedure Volume"),
            "Collagen Powder Vol": row.get("Collagen Powder - Procedure Volume"),
            "Wound Care DME Vol": row.get("Wound Care DME - Procedure Volume"),
            "All DME Vol": row.get("All of DME - Procedure Volume"),
        }
    return lookup


def build_collagen_lookup(csv_path):
    """Build NPI → collagen data lookup from a pipeline CSV."""
    df = pd.read_csv(csv_path, low_memory=False)
    lookup = {}
    for _, row in df.iterrows():
        npi = str(row.get("HCP NPI", "")).strip()
        if not npi or npi == "nan":
            continue
        lookup[npi] = {
            "Lg Collagen Vol": row.get("Lg Collagen Vol"),
            "Sm/Md Collagen Vol": row.get("Sm/Md Collagen Vol"),
            "Collagen Powder Vol": row.get("Collagen Powder Vol"),
            "Lg Incision Likelihood": row.get("Lg Incision Likelihood"),
        }
    return lookup


def num_or_none(v):
    """Convert to float, or return None."""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    try:
        f = float(v)
        if pd.isna(f):
            return None
        return f
    except (ValueError, TypeError):
        return None


def add_columns_to_tracker(wb, tab_name, lookup, has_incision=True):
    """Insert collagen/incision columns into a Call Tracker tab, after the Vol column."""
    ws = wb[tab_name]
    headers = [c.value for c in list(ws.iter_rows(max_row=1))[0]]

    # Find existing volume column index (Q = col 17)
    vol_col_idx = None
    for i, h in enumerate(headers, 1):
        if h in ("Joint Repl Vol", "Open Spine Vol", "Procedure Vol"):
            vol_col_idx = i
            break
    if vol_col_idx is None:
        logger.warning(f"No volume column found in {tab_name}")
        return

    # Columns to add after vol_col_idx
    new_cols = ["Lg Collagen Vol", "Sm/Md Collagen Vol", "Collagen Powder Vol"]
    if has_incision:
        new_cols.append("Lg Incision Likelihood")

    # Insert empty columns after vol_col_idx
    insert_at = vol_col_idx + 1
    for i in range(len(new_cols)):
        ws.insert_cols(insert_at)

    # Write new headers
    for i, col_name in enumerate(new_cols):
        col_idx = insert_at + i
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_NAVY
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Build NPI → row index map
    npi_col_idx = 1  # A
    npi_to_row = {}
    for row_idx in range(2, ws.max_row + 1):
        npi = ws.cell(row=row_idx, column=npi_col_idx).value
        if npi:
            npi_to_row[str(npi)] = row_idx

    # Fill in values
    filled_count = 0
    incision_counts = {}
    for npi, data in lookup.items():
        if npi not in npi_to_row:
            continue
        row_idx = npi_to_row[npi]
        is_alt = (row_idx - 2) % 2 == 1
        base_fill = ALT_ROW if is_alt else WHITE_FILL

        for i, col_name in enumerate(new_cols):
            col_idx = insert_at + i
            val = data.get(col_name)
            if col_name == "Lg Incision Likelihood":
                if val and isinstance(val, str) and val.strip() and val.strip() != "nan":
                    val = val.strip()
                    incision_counts[val] = incision_counts.get(val, 0) + 1
                else:
                    val = None
            else:
                val = num_or_none(val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = BORDER
            cell.fill = base_fill

            # Color-code incision likelihood
            if col_name == "Lg Incision Likelihood" and val in INCISION_FILLS:
                cell.fill = INCISION_FILLS[val]
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        filled_count += 1

    # Update auto_filter range
    num_cols = len(headers) + len(new_cols)
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"

    logger.info(f"  {tab_name}: +{len(new_cols)} cols, filled {filled_count}/{len(npi_to_row)} rows")
    if incision_counts:
        logger.info(f"    Incision distribution: {incision_counts}")
    return new_cols, insert_at


def compute_incision_for_jr(jr_tab):
    """Compute Lg Incision Likelihood for JR leads based on Joint Repl Vol + Open Ortho Vol.

    Heuristic that matches the Spine/Outside categorization logic:
      - High: Joint Repl Vol > 500 or Open Ortho Vol > 200
      - Medium-High: Joint Repl Vol 200-500 or Open Ortho Vol 50-200
      - Medium: Joint Repl Vol 50-200
      - Low: Joint Repl Vol 10-50
      - Unlikely: Joint Repl Vol < 10 (non-surgical or very low)
    """
    # We'll compute this in a post-pass when writing the column
    pass


def main():
    input_file = "Master_Lead_List_Tracker (3).xlsx"

    logger.info("Loading workbook...")
    wb = openpyxl.load_workbook(input_file)

    # Build lookups
    logger.info("Building collagen lookups...")
    jr_lookup = build_collagen_lookup_jr()
    sn_lookup = build_collagen_lookup("data/output/csv_export/Spine_Neuro.csv")
    oos_lookup = build_collagen_lookup("data/output/csv_export/Outside_Ortho_Spine.csv")
    logger.info(f"  JR: {len(jr_lookup)} NPIs")
    logger.info(f"  S&N: {len(sn_lookup)} NPIs")
    logger.info(f"  OOS: {len(oos_lookup)} NPIs")

    # Compute Incision Likelihood for JR based on volumes
    jr_csv = pd.read_csv("data/input/joint_replacement_targeting.csv", dtype=str)
    jr_csv.columns = [c.strip('"') for c in jr_csv.columns]
    for _, row in jr_csv.iterrows():
        npi = str(row.get("HCP NPI", "")).strip()
        if not npi or npi not in jr_lookup:
            continue
        joint_vol = num_or_none(row.get("Joint Replacement - Procedure Volume")) or 0
        open_ortho = num_or_none(row.get("Open Orthopedic Procedures - Procedure Volume")) or 0

        # JR heuristic for incision likelihood
        if joint_vol > 500 or open_ortho > 200:
            likelihood = "High"
        elif joint_vol >= 200 or open_ortho >= 50:
            likelihood = "Medium-High"
        elif joint_vol >= 50:
            likelihood = "Medium"
        elif joint_vol >= 10:
            likelihood = "Low"
        else:
            likelihood = "Unlikely"
        jr_lookup[npi]["Lg Incision Likelihood"] = likelihood

    # Check if collagen cols already added (idempotent run)
    for tab in ["Call Tracker - JR", "Call Tracker - S&N", "Call Tracker - OOS"]:
        ws = wb[tab]
        headers = [c.value for c in list(ws.iter_rows(max_row=1))[0]]
        if "Lg Collagen Vol" in headers:
            logger.warning(f"{tab} already has collagen columns — skipping insertion")
            continue

    # Add columns to each Call Tracker
    logger.info("\nAdding collagen/incision columns to Call Trackers...")
    add_columns_to_tracker(wb, "Call Tracker - JR", jr_lookup, has_incision=True)
    add_columns_to_tracker(wb, "Call Tracker - S&N", sn_lookup, has_incision=True)
    add_columns_to_tracker(wb, "Call Tracker - OOS", oos_lookup, has_incision=True)

    wb.save(input_file)
    logger.info(f"Saved to {input_file}")


if __name__ == "__main__":
    main()
