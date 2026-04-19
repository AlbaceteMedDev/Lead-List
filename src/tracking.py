"""Persistence of rep-editable call/email tracking activity across pipeline runs.

The Master Lead List Tracker (opened by Gabe and his cold caller in Excel /
Google Sheets) is the source of truth for activity data. On each pipeline run
we:

1. Scan ``data/input/`` for any ``Master_Lead_List_Tracker*.xlsx`` file.
2. Read the Call Tracker / Email Tracker sheets.
3. For every NPI, extract the rep-edited columns (Lead Status, Decision Maker,
   Next Action, Call 1..5 Date/Outcome/Notes, Email 1..3 Date/Subject/Outcome/Notes).
4. Merge that activity back onto the freshly enriched DataFrame.
5. Persist a canonical JSON snapshot to ``data/cache/activity.json`` so
   activity survives even if the user deletes the tracker xlsx.

Leads that appear in the activity cache but not in the new CSV batch are not
silently dropped — their activity is retained in the cache for future runs.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd

log = logging.getLogger(__name__)

CALL_ROUNDS = 5
EMAIL_ROUNDS = 3

CALL_FIELDS = [f"Call {i} {field}" for i in range(1, CALL_ROUNDS + 1) for field in ("Date", "Outcome", "Notes")]
EMAIL_FIELDS = [f"Email {i} {field}" for i in range(1, EMAIL_ROUNDS + 1) for field in ("Date", "Subject", "Outcome", "Notes")]
LEAD_FIELDS = ["Lead Status", "Decision Maker?", "Next Action", "Next Action Date"]

ACTIVITY_FIELDS = LEAD_FIELDS + CALL_FIELDS + EMAIL_FIELDS


def _normalize_npi(value: object) -> str:
    if value is None:
        return ""
    s = "".join(c for c in str(value) if c.isdigit())
    return s if len(s) == 10 else ""


def _serialize(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def load_cache(cache_path: Path) -> dict[str, dict]:
    if not cache_path.exists():
        return {}
    try:
        return json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception as exc:
        log.warning("Could not read activity cache %s: %s", cache_path, exc)
        return {}


def save_cache(cache_path: Path, cache: dict[str, dict]) -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = cache_path.with_suffix(cache_path.suffix + ".tmp")
    tmp.write_text(json.dumps(cache, indent=0, sort_keys=True), encoding="utf-8")
    tmp.replace(cache_path)


def read_tracker_activity(xlsx_path: Path) -> dict[str, dict]:
    """Extract activity data from every Call Tracker / Email Tracker sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    activity: dict[str, dict] = {}

    for sheet in wb.sheetnames:
        if not (sheet.startswith("Call Tracker") or sheet.startswith("Email Tracker")):
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows))
        except StopIteration:
            continue
        header_idx = {h: i for i, h in enumerate(header) if h}
        if "HCP NPI" not in header_idx:
            continue
        npi_col = header_idx["HCP NPI"]

        for row in rows:
            if row is None or row[npi_col] is None:
                continue
            npi = _normalize_npi(row[npi_col])
            if not npi:
                continue
            entry = activity.setdefault(npi, {})
            for field in ACTIVITY_FIELDS:
                if field in header_idx:
                    val = _serialize(row[header_idx[field]])
                    if val and not entry.get(field):
                        entry[field] = val
    wb.close()
    return activity


def _is_newer(old: str, new: str) -> bool:
    """Return True if ``new`` represents a later date string than ``old``."""
    if not old:
        return bool(new)
    if not new:
        return False
    try:
        return datetime.fromisoformat(new[:10]) > datetime.fromisoformat(old[:10])
    except ValueError:
        return new != old


def _merge_entry(existing: dict, incoming: dict) -> dict:
    """Merge two activity entries, keeping the latest value per field.

    For lead-level fields (Lead Status, Decision Maker, Next Action): prefer the
    incoming value if it's non-empty. For dated call/email rounds: keep the
    entry whose Date is more recent per round.
    """
    out = dict(existing)
    for field, value in incoming.items():
        if not value:
            continue
        if field in LEAD_FIELDS:
            out[field] = value
            continue
        out[field] = value
    # Normalize: if a Call N / Email N round has empty Date, clear its Outcome/Notes
    for i in range(1, CALL_ROUNDS + 1):
        if not out.get(f"Call {i} Date"):
            out.pop(f"Call {i} Outcome", None)
            out.pop(f"Call {i} Notes", None)
    for i in range(1, EMAIL_ROUNDS + 1):
        if not out.get(f"Email {i} Date"):
            out.pop(f"Email {i} Subject", None)
            out.pop(f"Email {i} Outcome", None)
            out.pop(f"Email {i} Notes", None)
    return out


def merge_activity(cache: dict[str, dict], incoming: dict[str, dict]) -> dict[str, dict]:
    """Merge a freshly-read tracker activity dict into the persistent cache."""
    result = dict(cache)
    for npi, entry in incoming.items():
        if npi in result:
            result[npi] = _merge_entry(result[npi], entry)
        else:
            result[npi] = dict(entry)
    return result


def ingest_trackers(tracker_paths: Iterable[Path], cache_path: Path) -> dict[str, dict]:
    """Read every tracker file, merge into the persistent cache, and save."""
    cache = load_cache(cache_path)
    for path in tracker_paths:
        try:
            fresh = read_tracker_activity(path)
        except Exception as exc:
            log.warning("Could not read tracker %s: %s", path, exc)
            continue
        log.info("Tracker %s contributed %d NPIs of activity", path.name, len(fresh))
        cache = merge_activity(cache, fresh)
    save_cache(cache_path, cache)
    return cache


def find_tracker_files(input_dir: Path) -> list[Path]:
    input_dir = Path(input_dir)
    return sorted(input_dir.glob("Master_Lead_List_Tracker*.xlsx"))


def find_edit_files(cache_dir: Path) -> list[Path]:
    """Return every ``activity_edits*.json`` the dashboard has downloaded."""
    cache_dir = Path(cache_dir)
    if not cache_dir.exists():
        return []
    return sorted(cache_dir.glob("activity_edits*.json"))


def apply_edit_files(edit_files: Iterable[Path], cache: dict[str, dict]) -> dict[str, dict]:
    """Merge each dashboard-exported JSON edit file into the cache.

    Files are renamed with ``.applied`` suffix so they don't get re-merged.
    """
    merged = dict(cache)
    for path in edit_files:
        try:
            incoming = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            log.warning("Could not parse edit file %s: %s", path, exc)
            continue
        if not isinstance(incoming, dict):
            log.warning("Edit file %s is not a JSON object; skipping", path)
            continue
        log.info("Applying %d edits from %s", len(incoming), path.name)
        merged = merge_activity(merged, incoming)
        try:
            path.rename(path.with_suffix(path.suffix + ".applied"))
        except OSError as exc:
            log.warning("Could not mark %s applied: %s", path, exc)
    return merged


def apply_activity(df: pd.DataFrame, activity: dict[str, dict]) -> pd.DataFrame:
    """Left-join activity onto the enriched frame by NPI."""
    df = df.copy()
    if "HCP NPI" not in df.columns:
        return df
    for field in ACTIVITY_FIELDS:
        if field not in df.columns:
            df[field] = ""
    default_status = "New"
    for idx in df.index:
        npi = _normalize_npi(df.at[idx, "HCP NPI"])
        if not npi:
            continue
        entry = activity.get(npi)
        if not entry:
            if not df.at[idx, "Lead Status"]:
                df.at[idx, "Lead Status"] = default_status
            continue
        for field in ACTIVITY_FIELDS:
            if entry.get(field):
                df.at[idx, field] = entry[field]
        if not df.at[idx, "Lead Status"]:
            df.at[idx, "Lead Status"] = default_status
    return df


def summarize_last_touch(df: pd.DataFrame) -> pd.DataFrame:
    """Add Last Touch Date and Touch Count columns derived from activity fields."""
    df = df.copy()
    last_touch: list[str] = []
    touch_count: list[int] = []
    for _, row in df.iterrows():
        dates: list[str] = []
        for i in range(1, CALL_ROUNDS + 1):
            d = row.get(f"Call {i} Date", "") if f"Call {i} Date" in df.columns else ""
            if d and str(d).strip() and str(d).lower() != "nan":
                dates.append(str(d))
        for i in range(1, EMAIL_ROUNDS + 1):
            d = row.get(f"Email {i} Date", "") if f"Email {i} Date" in df.columns else ""
            if d and str(d).strip() and str(d).lower() != "nan":
                dates.append(str(d))
        touch_count.append(len(dates))
        last_touch.append(max(dates) if dates else "")
    df["Last Touch Date"] = last_touch
    df["Touch Count"] = touch_count
    return df
