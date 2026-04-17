"""Score each lead and surface Top Targets across all three portfolios.

Target Score (0-100) composite signals:
  + Practice Type: Private Practice=25pts, Hospital=0pts (easier to reach DM)
  + Tier: T1=20, T2=15, T3=10, T4=5, T5-6=0, Hospital=5
  + Microlyte Eligibility: Yes=15, No=5 (both get ProPacks)
  + Volume percentile within portfolio: top 10%=20, top 25%=15, top 50%=8
  + Incision Likelihood: High=15, Medium-High=10, Medium=5, Low=2, Unlikely=0
  + Greenfield (high vol + NO collagen vendor): +15 (no competitor to displace)
  + No collagen vendor (any vol): +5
  + Heavy collagen buyer (100+): -5 (competitor risk — already has a vendor)
  + Verified contact info: +5 (can actually reach them)

Total possible: ~115 → cap at 100.

Target Tier Flag:
  A+ (90+): Apex targets — Private Practice + Tier 1-2 + High incision + volume + collagen user
  A  (75-89): Top priority
  B  (60-74): Strong targets
  C  (40-59): Secondary
  D  (<40): Deprioritize
"""

import logging
import re
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("target_score")

# Colors
NAVY = "1F4E78"
PURPLE = "7030A0"
RED = "C00000"
DARK_GRAY = "3B3B3B"

# Target tier fills
AP_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # dark green
A_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")    # light green
B_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")    # amber
C_FILL = PatternFill(start_color="F79646", end_color="F79646", fill_type="solid")    # orange
D_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")    # red

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
WHITE_BOLD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

ALT_ROW = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, float) and pd.isna(v):
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def format_phone(phone):
    if not phone:
        return None
    digits = re.sub(r"\D", "", str(phone))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
    return phone


def compute_score(row, vol_col, vol_p75, vol_p90, vol_p50):
    """Compute target score 0-100 + reasons list."""
    reasons = []
    score = 0

    # Practice Type — private practices are easier to close
    ptype = str(row.get("Practice Type", "") or "")
    if ptype == "Private Practice":
        score += 25
        reasons.append("Private Practice")

    # Drive-time tier
    tier = str(row.get("Tier", "") or "")
    if "Tier 1" in tier:
        score += 20
        reasons.append("Tier 1 drive")
    elif "Tier 2" in tier:
        score += 15
        reasons.append("Tier 2 drive")
    elif "Tier 3" in tier:
        score += 10
    elif "Tier 4" in tier:
        score += 5
    elif tier == "Hospital-Based":
        score += 5

    # Microlyte eligibility
    microlyte = str(row.get("Microlyte Eligible", "") or "")
    if microlyte == "Yes":
        score += 15
        reasons.append("Microlyte eligible")
    else:
        score += 5  # still a ProPacks candidate

    # Volume percentile
    vol = safe_float(row.get(vol_col))
    if vol >= vol_p90:
        score += 20
        reasons.append("Top 10% volume")
    elif vol >= vol_p75:
        score += 15
        reasons.append("Top 25% volume")
    elif vol >= vol_p50:
        score += 8

    # Incision likelihood
    inc = str(row.get("Lg Incision Likelihood", "") or "")
    if inc == "High":
        score += 15
        reasons.append("High incision likelihood")
    elif inc == "Medium-High":
        score += 10
        reasons.append("Med-High incision")
    elif inc == "Medium":
        score += 5
    elif inc == "Low":
        score += 2

    # Collagen usage — GREENFIELD vs COMPETITOR
    # High volume + NO collagen = best target (no competitor to displace)
    # High volume + heavy collagen = they already have a vendor (harder sell)
    lg = safe_float(row.get("Lg Collagen Vol"))
    smmd = safe_float(row.get("Sm/Md Collagen Vol"))
    pwd = safe_float(row.get("Collagen Powder Vol"))
    total_collagen = lg + smmd + pwd

    if total_collagen == 0 and vol >= vol_p50:
        score += 15
        reasons.append("Greenfield — high volume, no wound care vendor")
    elif total_collagen == 0:
        score += 5
        reasons.append("No current wound care vendor")
    elif total_collagen >= 100:
        score -= 5
        reasons.append(f"⚠ Competitor risk — heavy collagen buyer ({int(total_collagen)})")
    elif total_collagen > 0:
        reasons.append(f"Light collagen user ({int(total_collagen)}) — switchable")

    # Verified contact info
    phone_status = str(row.get("Phone Status", "") or "")
    email_status = str(row.get("Email Status", "") or "")
    has_phone = phone_status in ("Verified", "Added from NPPES", "Updated (NPPES differs)")
    has_email = "Missing" not in email_status and email_status != ""
    if has_phone and has_email:
        score += 5
        # don't add to reasons — this is baseline hygiene
    elif not has_phone and not has_email:
        reasons.append("⚠ No verified contact")

    return min(score, 100), reasons


def classify_target_tier(score):
    """Classify composite score into a letter tier."""
    if score >= 90:
        return "A+"
    elif score >= 75:
        return "A"
    elif score >= 60:
        return "B"
    elif score >= 40:
        return "C"
    else:
        return "D"


def best_approach(row, tier_letter):
    """Recommend the best outreach approach based on signals."""
    tier = str(row.get("Tier", "") or "")
    ptype = str(row.get("Practice Type", "") or "")
    phone_status = str(row.get("Phone Status", "") or "")
    email_status = str(row.get("Email Status", "") or "")

    # Tier 1-2 Private Practice with High/Med-High incision → In-Person Lunch & Learn
    if tier_letter in ("A+", "A") and ("Tier 1" in tier or "Tier 2" in tier) and ptype == "Private Practice":
        return "In-Person Lunch & Learn"
    # High score but further away → Video call first
    if tier_letter in ("A+", "A") and ("Tier 3" in tier or "Tier 4" in tier):
        return "Video Call → In-Person"
    # Verified email but no phone, Tier 5+ → Email first
    if "Missing" not in email_status and phone_status in ("Missing", ""):
        return "Email First"
    # Phone verified → Cold Call
    if phone_status in ("Verified", "Added from NPPES"):
        if tier_letter in ("A+", "A", "B"):
            return "Cold Call"
        return "Email First"
    # Missing both → deprioritize
    if "Missing" in email_status and phone_status in ("Missing", ""):
        return "Deprioritize / Manual Research"
    return "Email First"


def score_tab(wb, tab_name, vol_col):
    """Add Target Score, Target Tier, Why, Best Approach columns to a Call Tracker."""
    ws = wb[tab_name]
    headers = [c.value for c in list(ws.iter_rows(max_row=1))[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    df = pd.DataFrame(rows, columns=headers)

    # Compute volume percentiles
    vols = pd.to_numeric(df[vol_col], errors="coerce").dropna()
    vols = vols[vols > 0]
    if len(vols) > 0:
        p50 = vols.quantile(0.5)
        p75 = vols.quantile(0.75)
        p90 = vols.quantile(0.9)
    else:
        p50 = p75 = p90 = 0

    logger.info(f"  {tab_name}: vol p50={p50:.0f}, p75={p75:.0f}, p90={p90:.0f}")

    # Score each row
    scores = []
    tiers = []
    whys = []
    approaches = []
    for _, row in df.iterrows():
        score, reasons = compute_score(row, vol_col, p75, p90, p50)
        tier = classify_target_tier(score)
        scores.append(score)
        tiers.append(tier)
        whys.append(" • ".join(reasons) if reasons else "—")
        approaches.append(best_approach(row, tier))

    df["Target Score"] = scores
    df["Target Tier"] = tiers
    df["Why Target?"] = whys
    df["Best Approach"] = approaches

    # Check if columns already exist (idempotent)
    existing = {h for h in headers}
    new_cols = [c for c in ["Target Score", "Target Tier", "Why Target?", "Best Approach"]
                if c not in existing]

    if not new_cols:
        logger.info(f"  {tab_name}: scoring columns already exist, updating values")
        # Update in place
        col_idx_map = {h: headers.index(h) + 1 for h in ["Target Score", "Target Tier", "Why Target?", "Best Approach"]}
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_name, value in [
                ("Target Score", row["Target Score"]),
                ("Target Tier", row["Target Tier"]),
                ("Why Target?", row["Why Target?"]),
                ("Best Approach", row["Best Approach"]),
            ]:
                if col_name in col_idx_map:
                    ws.cell(row=row_idx, column=col_idx_map[col_name], value=value)
    else:
        # Insert new columns at the end
        insert_at = len(headers) + 1
        for i, col_name in enumerate(new_cols):
            col_idx = insert_at + i
            header_cell = ws.cell(row=1, column=col_idx, value=col_name)
            header_cell.fill = HEADER_FILL
            header_cell.font = HEADER_FONT
            header_cell.alignment = CENTER
            header_cell.border = BORDER
            # Set widths
            if col_name == "Target Score":
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            elif col_name == "Target Tier":
                ws.column_dimensions[get_column_letter(col_idx)].width = 11
            elif col_name == "Why Target?":
                ws.column_dimensions[get_column_letter(col_idx)].width = 40
            elif col_name == "Best Approach":
                ws.column_dimensions[get_column_letter(col_idx)].width = 22

        # Write values
        score_idx = insert_at + new_cols.index("Target Score") if "Target Score" in new_cols else None
        tier_idx = insert_at + new_cols.index("Target Tier") if "Target Tier" in new_cols else None
        why_idx = insert_at + new_cols.index("Why Target?") if "Why Target?" in new_cols else None
        approach_idx = insert_at + new_cols.index("Best Approach") if "Best Approach" in new_cols else None

        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            is_alt = (row_idx - 2) % 2 == 1
            base = ALT_ROW if is_alt else WHITE_FILL

            if score_idx:
                c = ws.cell(row=row_idx, column=score_idx, value=row["Target Score"])
                c.font = BOLD_FONT
                c.alignment = CENTER
                c.fill = base
                c.border = BORDER
            if tier_idx:
                c = ws.cell(row=row_idx, column=tier_idx, value=row["Target Tier"])
                c.font = WHITE_BOLD
                c.alignment = CENTER
                c.border = BORDER
                fills = {"A+": AP_FILL, "A": A_FILL, "B": B_FILL, "C": C_FILL, "D": D_FILL}
                c.fill = fills.get(row["Target Tier"], base)
            if why_idx:
                c = ws.cell(row=row_idx, column=why_idx, value=row["Why Target?"])
                c.font = DATA_FONT
                c.alignment = LEFT
                c.fill = base
                c.border = BORDER
            if approach_idx:
                c = ws.cell(row=row_idx, column=approach_idx, value=row["Best Approach"])
                c.font = DATA_FONT
                c.alignment = CENTER
                c.fill = base
                c.border = BORDER

        # Update auto_filter
        total_cols = len(headers) + len(new_cols)
        last_row = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{last_row}"

    # Return the scored df for building Top Targets tabs
    return df


def build_top_targets_tab(wb, tab_name, df, vol_col, accent_color, top_n=250):
    """Build a Top Targets tab with the highest-scoring leads."""
    # Delete if exists
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(tab_name)

    # Sort by Target Score descending, then by volume descending
    df = df.copy()
    df["_sort_vol"] = pd.to_numeric(df[vol_col], errors="coerce").fillna(0)
    df = df.sort_values(["Target Score", "_sort_vol"], ascending=[False, False])
    df = df.head(top_n)
    df = df.drop(columns=["_sort_vol"])

    # Columns to include
    display_cols = [
        "Target Tier", "Target Score",
        "HCP NPI", "First Name", "Last Name", "Credential", "Specialty",
        "Phone", "Email", "Primary Site of Care", "City", "State",
        "Practice Type", "Tier", "MAC Jurisdiction", "Microlyte Eligible",
        vol_col, "Lg Collagen Vol", "Sm/Md Collagen Vol", "Collagen Powder Vol",
        "Lg Incision Likelihood",
        "Why Target?", "Best Approach",
        "Lead Status", "Next Action", "Next Action Date", "Decision Maker?", "Notes",
    ]
    # Filter to existing columns
    display_cols = [c for c in display_cols if c in df.columns]

    # Column widths
    widths = {
        "Target Tier": 10, "Target Score": 11,
        "HCP NPI": 14, "First Name": 13, "Last Name": 15, "Credential": 10,
        "Specialty": 28, "Phone": 16, "Email": 32,
        "Primary Site of Care": 32, "City": 15, "State": 8,
        "Practice Type": 16, "Tier": 20, "MAC Jurisdiction": 16, "Microlyte Eligible": 14,
        vol_col: 14, "Lg Collagen Vol": 14, "Sm/Md Collagen Vol": 14, "Collagen Powder Vol": 14,
        "Lg Incision Likelihood": 16,
        "Why Target?": 45, "Best Approach": 22,
        "Lead Status": 16, "Next Action": 20, "Next Action Date": 14,
        "Decision Maker?": 14, "Notes": 30,
    }

    # Add title row (merged)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_cols))
    title_cell = ws.cell(row=1, column=1, value=f"  ⭐ TOP {len(df)} TARGETS — {tab_name.replace('Top Targets - ', '')}")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=accent_color, end_color=accent_color, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # Header row (row 2)
    for i, h in enumerate(display_cols, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 15)
    ws.row_dimensions[2].height = 32

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        is_alt = (row_idx - 3) % 2 == 1
        base = ALT_ROW if is_alt else WHITE_FILL

        for col_idx, h in enumerate(display_cols, 1):
            val = row.get(h)
            if isinstance(val, float) and pd.isna(val):
                val = None
            if h == "Phone" and val:
                val = format_phone(val)
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = DATA_FONT
            c.border = BORDER
            c.fill = base
            if h == "Why Target?":
                c.alignment = LEFT
            else:
                c.alignment = CENTER

            # Color-code Target Tier
            if h == "Target Tier":
                tier = row.get("Target Tier")
                fills = {"A+": AP_FILL, "A": A_FILL, "B": B_FILL, "C": C_FILL, "D": D_FILL}
                if tier in fills:
                    c.fill = fills[tier]
                    c.font = WHITE_BOLD

            # Color-code incision likelihood
            if h == "Lg Incision Likelihood" and val:
                inc_fills = {
                    "High": AP_FILL, "Medium-High": A_FILL, "Medium": B_FILL,
                    "Low": C_FILL, "Unlikely": D_FILL,
                }
                if val in inc_fills:
                    c.fill = inc_fills[val]
                    c.font = WHITE_BOLD

    num_rows = len(df) + 2
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_cols))}{num_rows}"

    # Add Lead Status dropdown
    try:
        ls_idx = display_cols.index("Lead Status") + 1
        col_l = get_column_letter(ls_idx)
        dv = DataValidation(type="list", formula1='"New,Contacted,Meeting Scheduled,Meeting Completed,Proposal Sent,Won,Lost,Nurture"', allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_l}3:{col_l}{num_rows}")
    except ValueError:
        pass

    # Add Decision Maker dropdown
    try:
        dm_idx = display_cols.index("Decision Maker?") + 1
        col_l = get_column_letter(dm_idx)
        dv = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_l}3:{col_l}{num_rows}")
    except ValueError:
        pass

    logger.info(f"  Built {tab_name} with {len(df)} top leads")
    return df


def main():
    input_file = "Master_Lead_List_Tracker (3).xlsx"

    logger.info("Loading workbook...")
    wb = openpyxl.load_workbook(input_file)

    logger.info("\nScoring Call Tracker - JR...")
    jr_df = score_tab(wb, "Call Tracker - JR", "Joint Repl Vol")

    logger.info("\nScoring Call Tracker - S&N...")
    sn_df = score_tab(wb, "Call Tracker - S&N", "Open Spine Vol")

    logger.info("\nScoring Call Tracker - OOS...")
    oos_df = score_tab(wb, "Call Tracker - OOS", "Procedure Vol")

    # Log tier distributions
    for name, df in [("JR", jr_df), ("S&N", sn_df), ("OOS", oos_df)]:
        counts = df["Target Tier"].value_counts()
        logger.info(f"\n{name} target tier distribution:")
        for t in ["A+", "A", "B", "C", "D"]:
            logger.info(f"  {t}: {counts.get(t, 0)}")

    # Build Top Targets tabs
    logger.info("\nBuilding Top Targets tabs...")
    build_top_targets_tab(wb, "Top Targets - JR", jr_df, "Joint Repl Vol", NAVY)
    build_top_targets_tab(wb, "Top Targets - S&N", sn_df, "Open Spine Vol", PURPLE)
    build_top_targets_tab(wb, "Top Targets - OOS", oos_df, "Procedure Vol", RED)

    # Reorder tabs: Dashboard | Top Targets | JR | S&N | OOS
    desired_order = [
        "Dashboard",
        "Top Targets - JR", "Top Targets - S&N", "Top Targets - OOS",
        "Call Tracker - JR", "Email Tracker - JR", "Email Drafts - JR",
        "Call Tracker - S&N", "Email Tracker - S&N", "Email Drafts - S&N",
        "Call Tracker - OOS", "Email Tracker - OOS", "Email Drafts - OOS",
    ]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            if cur != i:
                wb.move_sheet(name, offset=i - cur)

    wb.save(input_file)
    logger.info(f"\nSaved to {input_file}")
    logger.info(f"Final tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
