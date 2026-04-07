"""Generate the Cold Calling & Outreach Tracker workbook (6 tabs)."""

import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Arial", size=9)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Conditional formatting fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")


def _style_header(ws, num_cols):
    """Apply header styling to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _auto_filter(ws, num_cols, num_rows):
    """Apply auto-filter."""
    if num_rows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"


def _add_dropdown(ws, col_letter, options, min_row=2, max_row=5000):
    """Add data validation dropdown to a column."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please select from the dropdown list"
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{min_row}:{col_letter}{max_row}")


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _compute_priority(row):
    """
    A: Private Practice + Tier 1-2 + top 25% volume
    B: Private Practice + Tier 3-4, OR any tier with moderate volume (25th-75th pctl)
    C: Everything else — Tier 5-6, hospital-based, or bottom 25% volume
    """
    ptype = str(row.get("Practice Type", ""))
    tier = str(row.get("Tier", ""))
    vol = row.get("_total_vol", 0)
    p75 = row.get("_p75", 0)
    p25 = row.get("_p25", 0)

    is_private = ptype == "Private Practice"
    is_tier_12 = "Tier 1" in tier or "Tier 2" in tier
    is_tier_34 = "Tier 3" in tier or "Tier 4" in tier
    is_tier_56 = "Tier 5" in tier or "Tier 6" in tier
    is_hospital = ptype == "Hospital-Based"

    if is_private and is_tier_12 and vol >= p75:
        return "A"
    if is_private and is_tier_34:
        return "B"
    if is_private and is_tier_12 and vol >= p25:
        return "B"
    if vol >= p25 and vol < p75:
        return "B"
    return "C"


def build_tracker(input_xlsx: str, output_path: str = "data/output/outreach_tracker.xlsx"):
    """Build the full 6-tab outreach tracker workbook."""
    import openpyxl

    logger.info("Loading enriched lead data...")
    src_wb = openpyxl.load_workbook(input_xlsx, read_only=True)

    # Collect all leads from all tier tabs
    all_rows = []
    headers = None
    for name in src_wb.sheetnames:
        if name == "Summary":
            continue
        ws = src_wb[name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                if headers is None:
                    headers = list(row)
                continue
            all_rows.append(dict(zip(headers, row)))
    src_wb.close()

    logger.info(f"Loaded {len(all_rows)} leads")

    # Compute total volume and percentiles for priority scoring
    for row in all_rows:
        joint_vol = _safe_float(row.get("Joint Repl Vol"))
        # collagen volumes aren't in the output workbook, use what we have
        row["_total_vol"] = joint_vol

    vols = [r["_total_vol"] for r in all_rows if r["_total_vol"] > 0]
    if vols:
        vols_sorted = sorted(vols)
        p25 = vols_sorted[len(vols_sorted) // 4]
        p75 = vols_sorted[3 * len(vols_sorted) // 4]
    else:
        p25, p75 = 0, 0

    for row in all_rows:
        row["_p25"] = p25
        row["_p75"] = p75
        row["Lead Priority"] = _compute_priority(row)

    # Create workbook
    wb = Workbook()

    # =========================================================================
    # TAB 1: MASTER LEAD LIST
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Master Lead List"

    master_cols = [
        ("HCP NPI", 14), ("First Name", 13), ("Last Name", 15),
        ("Credential", 10), ("Specialty", 28),
        ("Verified Phone", 15), ("Phone Status", 16),
        ("Email", 32), ("Email Status", 22),
        ("Primary Site of Care", 32), ("Address 1", 25),
        ("City", 15), ("State", 8), ("Postal Code", 11),
        ("Practice Type", 16), ("Tier", 20),
        ("Lead Priority", 12), ("Lead Status", 18),
        ("MAC Jurisdiction", 16), ("Microlyte Eligible", 14),
        ("Joint Repl Vol", 12), ("Knee Vol", 10),
        ("Hip Vol", 10), ("Shoulder Vol", 12), ("Open Ortho Vol", 12),
        ("Medical School", 28), ("HCP URL", 30),
        ("Subject Line", 35), ("Draft Email", 60),
        ("Decision Maker?", 14), ("LinkedIn URL", 25),
        ("Practice Website", 25), ("Notes", 40),
    ]

    # Write headers
    for col_idx, (col_name, width) in enumerate(master_cols, start=1):
        ws1.cell(row=1, column=col_idx, value=col_name)
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    _style_header(ws1, len(master_cols))

    # Write data
    for row_idx, row in enumerate(all_rows, start=2):
        col_map = {
            "HCP NPI": row.get("HCP NPI"),
            "First Name": row.get("First Name"),
            "Last Name": row.get("Last Name"),
            "Credential": row.get("Credential"),
            "Specialty": row.get("Specialty"),
            "Verified Phone": row.get("Verified Phone"),
            "Phone Status": row.get("Phone Status"),
            "Email": row.get("Email"),
            "Email Status": row.get("Email Status"),
            "Primary Site of Care": row.get("Primary Site of Care"),
            "Address 1": row.get("Address 1"),
            "City": row.get("City"),
            "State": row.get("State"),
            "Postal Code": row.get("Postal Code"),
            "Practice Type": row.get("Practice Type"),
            "Tier": row.get("Tier"),
            "Lead Priority": row.get("Lead Priority"),
            "Lead Status": "New",
            "MAC Jurisdiction": row.get("MAC Jurisdiction"),
            "Microlyte Eligible": row.get("Microlyte Eligible"),
            "Joint Repl Vol": row.get("Joint Repl Vol"),
            "Knee Vol": row.get("Knee Vol"),
            "Hip Vol": row.get("Hip Vol"),
            "Shoulder Vol": row.get("Shoulder Vol"),
            "Open Ortho Vol": row.get("Open Ortho Vol"),
            "Medical School": row.get("Medical School"),
            "HCP URL": row.get("HCP URL"),
            "Subject Line": row.get("Subject Line"),
            "Draft Email": row.get("Draft Email"),
            "Decision Maker?": "Unknown",
            "LinkedIn URL": None,
            "Practice Website": None,
            "Notes": None,
        }
        for col_idx, (col_name, _) in enumerate(master_cols, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=col_map.get(col_name))
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

        # Conditional formatting: Lead Status
        status_col = 18  # Lead Status
        priority_col = 17  # Lead Priority
        status = col_map["Lead Status"]
        priority = col_map["Lead Priority"]

        if priority == "A":
            ws1.cell(row=row_idx, column=priority_col).fill = GREEN_FILL
        elif priority == "B":
            ws1.cell(row=row_idx, column=priority_col).fill = YELLOW_FILL
        elif priority == "C":
            ws1.cell(row=row_idx, column=priority_col).fill = RED_FILL

    num_rows = len(all_rows) + 1
    _auto_filter(ws1, len(master_cols), num_rows)

    # Dropdowns for Master Lead List
    _add_dropdown(ws1, "R", ["New", "Contacted", "Meeting Scheduled", "Meeting Completed",
                              "Proposal Sent", "Won", "Lost", "Nurture"], max_row=num_rows)
    _add_dropdown(ws1, "AD", ["Yes", "No", "Unknown"], max_row=num_rows)

    # =========================================================================
    # TAB 2: CALL LOG
    # =========================================================================
    ws2 = wb.create_sheet("Call Log")
    call_cols = [
        ("Date", 12), ("HCP NPI", 14), ("Lead Name", 22),
        ("Practice Name", 30), ("Phone Number Used", 16),
        ("Call Outcome", 28), ("Gatekeeper Name", 18),
        ("Spoke With (Name/Title)", 22), ("Best Time to Call", 16),
        ("Follow-Up Date", 14), ("Follow-Up Action", 22),
        ("Products Discussed", 20), ("Call Duration (min)", 14),
        ("Call Notes", 50),
    ]
    for col_idx, (col_name, width) in enumerate(call_cols, start=1):
        ws2.cell(row=1, column=col_idx, value=col_name)
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    _style_header(ws2, len(call_cols))
    _auto_filter(ws2, len(call_cols), 1)

    # Date format for column A
    ws2.column_dimensions["A"].number_format = "MM/DD/YYYY"

    _add_dropdown(ws2, "F", [
        "Connected - Meeting Set", "Connected - Call Back",
        "Connected - Not Interested", "Voicemail", "No Answer",
        "Wrong Number", "Gatekeeper - Message Left",
        "Gatekeeper - Blocked", "Disconnected",
    ])
    _add_dropdown(ws2, "K", [
        "Call Again", "Send Email", "In-Person Visit",
        "Send Materials", "Schedule Meeting", "No Follow-Up",
    ])

    # =========================================================================
    # TAB 3: EMAIL OUTREACH LOG
    # =========================================================================
    ws3 = wb.create_sheet("Email Outreach Log")
    email_cols = [
        ("Date Sent", 12), ("HCP NPI", 14), ("Lead Name", 22),
        ("Practice Name", 30), ("Email Used", 32),
        ("Email Type", 18), ("Subject Line", 35),
        ("Template Track", 14), ("Opened?", 10),
        ("Replied?", 10), ("Follow-Up Date", 14),
        ("Follow-Up Action", 22), ("Notes", 40),
    ]
    for col_idx, (col_name, width) in enumerate(email_cols, start=1):
        ws3.cell(row=1, column=col_idx, value=col_name)
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    _style_header(ws3, len(email_cols))
    _auto_filter(ws3, len(email_cols), 1)

    _add_dropdown(ws3, "F", [
        "Cold Intro", "Follow-Up", "Post-Call", "Post-Visit", "Proposal",
    ])
    _add_dropdown(ws3, "H", ["Track A (ProPacks)", "Track B (ProPacks + Microlyte)"])
    _add_dropdown(ws3, "I", ["Yes", "No", "Unknown"])
    _add_dropdown(ws3, "J", ["Yes", "No"])

    # =========================================================================
    # TAB 4: IN-PERSON VISIT LOG
    # =========================================================================
    ws4 = wb.create_sheet("In-Person Visit Log")
    visit_cols = [
        ("Date", 12), ("HCP NPI", 14), ("Lead Name", 22),
        ("Practice Name", 30), ("Address", 35),
        ("Visit Type", 22), ("Met With (Name/Title)", 25),
        ("Outcome", 28), ("Products Discussed", 22),
        ("Samples Left?", 12), ("Follow-Up Date", 14),
        ("Follow-Up Action", 22), ("Visit Notes", 50),
    ]
    for col_idx, (col_name, width) in enumerate(visit_cols, start=1):
        ws4.cell(row=1, column=col_idx, value=col_name)
        ws4.column_dimensions[get_column_letter(col_idx)].width = width
    _style_header(ws4, len(visit_cols))
    _auto_filter(ws4, len(visit_cols), 1)

    _add_dropdown(ws4, "F", [
        "Cold Drop-In", "Scheduled Meeting", "Follow-Up", "In-Service", "Lunch & Learn",
    ])
    _add_dropdown(ws4, "H", [
        "Positive - Next Steps", "Neutral - Left Materials",
        "Negative - Not Interested", "No One Available",
    ])
    _add_dropdown(ws4, "J", ["Yes", "No"])

    # =========================================================================
    # TAB 5: PIPELINE DASHBOARD DATA
    # =========================================================================
    ws5 = wb.create_sheet("Pipeline Dashboard")
    ws5.sheet_properties.tabColor = "1B4F72"

    title_font = Font(name="Arial", size=14, bold=True, color="1B4F72")
    section_font = Font(name="Arial", size=11, bold=True)
    data_font = Font(name="Arial", size=10)
    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 15
    ws5.column_dimensions["C"].width = 15

    row = 1
    ws5.cell(row=row, column=1, value="Pipeline Dashboard").font = title_font
    ws5.cell(row=row, column=2, value="Count").font = section_font
    ws5.cell(row=row, column=3, value="% of Total").font = section_font
    row += 1
    ws5.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y')}").font = data_font

    # Leads by Status
    row += 2
    ws5.cell(row=row, column=1, value="LEADS BY STATUS").font = section_font
    ws5.cell(row=row, column=1).fill = HEADER_FILL
    ws5.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    statuses = ["New", "Contacted", "Meeting Scheduled", "Meeting Completed",
                "Proposal Sent", "Won", "Lost", "Nurture"]
    for status in statuses:
        ws5.cell(row=row, column=1, value=status).font = data_font
        count = len(all_rows) if status == "New" else 0
        ws5.cell(row=row, column=2, value=count).font = data_font
        row += 1
    ws5.cell(row=row, column=1, value="TOTAL").font = section_font
    ws5.cell(row=row, column=2, value=len(all_rows)).font = section_font

    # Leads by Tier
    row += 2
    ws5.cell(row=row, column=1, value="LEADS BY TIER").font = section_font
    ws5.cell(row=row, column=1).fill = HEADER_FILL
    ws5.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    tier_counts = {}
    for r in all_rows:
        t = r.get("Tier", "Unknown")
        tier_counts[t] = tier_counts.get(t, 0) + 1
    for tier in ["Tier 1 (0-30 min)", "Tier 2 (30-60 min)", "Tier 3 (60-120 min)",
                 "Tier 4 (120-180 min)", "Tier 5 (180+ drivable)", "Tier 6 (Requires flight)",
                 "Hospital-Based"]:
        ws5.cell(row=row, column=1, value=tier).font = data_font
        ws5.cell(row=row, column=2, value=tier_counts.get(tier, 0)).font = data_font
        row += 1

    # Leads by Priority
    row += 1
    ws5.cell(row=row, column=1, value="LEADS BY PRIORITY").font = section_font
    ws5.cell(row=row, column=1).fill = HEADER_FILL
    ws5.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    priority_counts = {}
    for r in all_rows:
        p = r.get("Lead Priority", "C")
        priority_counts[p] = priority_counts.get(p, 0) + 1
    for p, label in [("A", "A (High Priority)"), ("B", "B (Medium)"), ("C", "C (Low)")]:
        ws5.cell(row=row, column=1, value=label).font = data_font
        ws5.cell(row=row, column=2, value=priority_counts.get(p, 0)).font = data_font
        fill = GREEN_FILL if p == "A" else YELLOW_FILL if p == "B" else RED_FILL
        ws5.cell(row=row, column=1).fill = fill
        row += 1

    # Microlyte eligibility
    row += 1
    ws5.cell(row=row, column=1, value="MICROLYTE ELIGIBILITY").font = section_font
    ws5.cell(row=row, column=1).fill = HEADER_FILL
    ws5.cell(row=row, column=2).fill = HEADER_FILL
    row += 1
    elig_yes = sum(1 for r in all_rows if r.get("Microlyte Eligible") == "Yes")
    elig_no = sum(1 for r in all_rows if r.get("Microlyte Eligible") == "No")
    ws5.cell(row=row, column=1, value="Eligible (Track B)").font = data_font
    ws5.cell(row=row, column=2, value=elig_yes).font = data_font
    row += 1
    ws5.cell(row=row, column=1, value="Not Eligible (Track A)").font = data_font
    ws5.cell(row=row, column=2, value=elig_no).font = data_font

    # Activity tracking placeholders
    row += 2
    ws5.cell(row=row, column=1, value="ACTIVITY TRACKING").font = section_font
    ws5.cell(row=row, column=1).fill = HEADER_FILL
    ws5.cell(row=row, column=2, value="This Week").font = section_font
    ws5.cell(row=row, column=2).fill = HEADER_FILL
    ws5.cell(row=row, column=3, value="This Month").font = section_font
    ws5.cell(row=row, column=3).fill = HEADER_FILL
    row += 1
    for metric in ["Calls Made", "Emails Sent", "Visits Completed",
                    "Meetings Set", "Proposals Sent"]:
        ws5.cell(row=row, column=1, value=metric).font = data_font
        ws5.cell(row=row, column=2, value=0).font = data_font
        ws5.cell(row=row, column=3, value=0).font = data_font
        row += 1

    # Conversion rates
    row += 1
    ws5.cell(row=row, column=1, value="CONVERSION RATES").font = section_font
    ws5.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for stage in ["Contacted → Meeting Set", "Meeting → Proposal",
                   "Proposal → Won", "Overall Win Rate"]:
        ws5.cell(row=row, column=1, value=stage).font = data_font
        ws5.cell(row=row, column=2, value="0%").font = data_font
        row += 1

    # =========================================================================
    # TAB 6: WEEKLY PLANNING
    # =========================================================================
    ws6 = wb.create_sheet("Weekly Planning")
    plan_cols = [
        ("Week Of", 14), ("Tier Focus", 20), ("Target Calls", 14),
        ("Target Emails", 14), ("Target Visits", 14),
        ("Actual Calls", 14), ("Actual Emails", 14), ("Actual Visits", 14),
        ("Priority Leads to Contact", 40), ("Notes / Strategy", 50),
    ]
    for col_idx, (col_name, width) in enumerate(plan_cols, start=1):
        ws6.cell(row=1, column=col_idx, value=col_name)
        ws6.column_dimensions[get_column_letter(col_idx)].width = width
    _style_header(ws6, len(plan_cols))
    _auto_filter(ws6, len(plan_cols), 1)

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info(f"Outreach tracker saved to {output_path}")
    return output_path


def export_csvs(input_xlsx: str, output_dir: str = "data/output/csv_export"):
    """Export each tab of the tracker as a separate CSV for Google Sheets import."""
    import openpyxl

    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.load_workbook(input_xlsx, read_only=True)

    for name in wb.sheetnames:
        ws = wb[name]
        safe_name = name.replace(" ", "_").replace("/", "-").replace("(", "").replace(")", "")
        csv_path = os.path.join(output_dir, f"{safe_name}.csv")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        df = pd.DataFrame(rows[1:], columns=rows[0])
        df.to_csv(csv_path, index=False)
        logger.info(f"  Exported {csv_path} ({len(df)} rows)")

    wb.close()
    logger.info(f"CSVs exported to {output_dir}/")


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    # Find latest enriched workbook
    output_dir = "data/output"
    xlsx_files = sorted([f for f in os.listdir(output_dir) if f.startswith("lead_list_enriched") and f.endswith(".xlsx")])
    if not xlsx_files:
        print("No enriched workbook found. Run the pipeline first (python run.py)")
        sys.exit(1)

    latest = os.path.join(output_dir, xlsx_files[-1])
    print(f"Using: {latest}")

    tracker_path = build_tracker(latest)
    export_csvs(tracker_path)
