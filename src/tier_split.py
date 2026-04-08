"""Split Spine & Neuro and Outside Ortho into tier-based tabs."""

import logging
import os
import sys

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("tier_split")

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Arial", size=9)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MICROLYTE_GREEN = PatternFill(start_color="D4EFDF", end_color="D4EFDF", fill_type="solid")
VERIFIED_GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
MISSING_RED = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
NPPES_YELLOW = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
ALT_ROW = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
HEADER_NAVY = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
HEADER_WHITE_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)

TIERS = [
    "Tier 1 (0-30 min)", "Tier 2 (30-60 min)", "Tier 3 (60-120 min)",
    "Tier 4 (120-180 min)", "Tier 5 (180+ drivable)", "Tier 6 (Requires flight)",
    "Hospital-Based",
]


def extract_tab(wb, sheet_name):
    """Extract a sheet into a list of dicts."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    return headers, [dict(zip(headers, r)) for r in rows[1:]]


def write_tier_sheet(wb, tab_name, headers, rows, col_widths, vol_col):
    """Write a single tier tab with formatting."""
    ws = wb.create_sheet(tab_name)

    # Write headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_NAVY
        cell.font = HEADER_WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Set column widths
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 15)

    # Find key column indices
    def col_idx_of(name):
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    priority_idx = col_idx_of("Lead Priority")
    email_status_idx = col_idx_of("Email Status")
    phone_status_idx = col_idx_of("Phone Status")
    microlyte_idx = col_idx_of("Microlyte Eligible")
    lead_status_idx = col_idx_of("Lead Status")

    # Sort by volume descending
    def sort_key(r):
        v = r.get(vol_col)
        if v is None:
            return 0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0
    rows = sorted(rows, key=sort_key, reverse=True)

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        is_alt = (row_idx - 2) % 2 == 1
        base_fill = ALT_ROW if is_alt else WHITE_FILL

        for col_idx, h in enumerate(headers, 1):
            val = row.get(h)
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = base_fill
            if h == "Draft Email":
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Priority coloring
        if priority_idx:
            p = row.get("Lead Priority", "")
            if p == "A":
                ws.cell(row=row_idx, column=priority_idx).fill = GREEN_FILL
            elif p == "B":
                ws.cell(row=row_idx, column=priority_idx).fill = YELLOW_FILL
            elif p == "C":
                ws.cell(row=row_idx, column=priority_idx).fill = RED_FILL

        # Email status coloring
        if email_status_idx:
            es = str(row.get("Email Status", ""))
            if "Verified" in es:
                ws.cell(row=row_idx, column=email_status_idx).fill = VERIFIED_GREEN
            elif es == "Missing":
                ws.cell(row=row_idx, column=email_status_idx).fill = MISSING_RED

        # Phone status coloring
        if phone_status_idx:
            ps = str(row.get("Phone Status", ""))
            if ps == "Added from NPPES":
                ws.cell(row=row_idx, column=phone_status_idx).fill = NPPES_YELLOW

        # Microlyte coloring
        if microlyte_idx:
            me = str(row.get("Microlyte Eligible", ""))
            if me == "Yes":
                ws.cell(row=row_idx, column=microlyte_idx).fill = MICROLYTE_GREEN

    num_rows = len(rows) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{num_rows}"

    # Dropdowns
    if lead_status_idx:
        col_l = get_column_letter(lead_status_idx)
        dv = DataValidation(type="list", formula1='"New,Contacted,Meeting Scheduled,Meeting Completed,Proposal Sent,Won,Lost,Nurture"', allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_l}2:{col_l}{num_rows}")

    for outcome_col in ["Call 1 Outcome", "Call 2 Outcome", "Call 3 Outcome"]:
        idx = col_idx_of(outcome_col)
        if idx:
            col_l = get_column_letter(idx)
            dv = DataValidation(type="list", formula1='"Connected - Meeting Set,Connected - Call Back,Connected - Not Interested,Voicemail,No Answer,Wrong Number,Gatekeeper - Message Left,Gatekeeper - Blocked"', allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_l}2:{col_l}{num_rows}")

    for outcome_col in ["Email 1 Outcome", "Email 2 Outcome"]:
        idx = col_idx_of(outcome_col)
        if idx:
            col_l = get_column_letter(idx)
            dv = DataValidation(type="list", formula1='"Opened,Replied,Bounced,No Response,Unsubscribed"', allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_l}2:{col_l}{num_rows}")

    dm_idx = col_idx_of("Decision Maker?")
    if dm_idx:
        col_l = get_column_letter(dm_idx)
        dv = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_l}2:{col_l}{num_rows}")

    logger.info(f"  {tab_name}: {len(rows)} leads")
    return ws


def main():
    input_file = "Master_Lead_List_Tracker (3).xlsx"

    logger.info("Loading workbook...")
    wb_read = openpyxl.load_workbook(input_file, read_only=True)

    # Extract both enriched tabs
    sn_headers, sn_rows = extract_tab(wb_read, "Spine & Neuro")
    oo_headers, oo_rows = extract_tab(wb_read, "Outside Ortho & Spine")
    wb_read.close()

    logger.info(f"Spine & Neuro: {len(sn_rows)} leads, {len(sn_headers)} cols")
    logger.info(f"Outside Ortho & Spine: {len(oo_rows)} leads, {len(oo_headers)} cols")

    # Open workbook in write mode
    wb = openpyxl.load_workbook(input_file)

    # Remove old flat tabs
    for name in ["Spine & Neuro", "Outside Ortho & Spine"]:
        if name in wb.sheetnames:
            del wb[name]

    # Column widths
    col_widths = {
        "HCP NPI": 14, "First Name": 13, "Last Name": 15, "Credential": 10,
        "Specialty": 28, "Verified Phone": 15, "Phone Status": 16,
        "Email": 32, "Email Status": 22, "Primary Site of Care": 32,
        "Address": 25, "City": 15, "State": 8, "Postal Code": 11,
        "Practice Type": 16, "Tier": 20, "Lead Priority": 12, "Lead Status": 16,
        "MAC Jurisdiction": 16, "Microlyte Eligible": 14,
        "Open Spine Vol": 14, "Procedure Vol": 14,
        "Lg Collagen Vol": 14, "Sm/Md Collagen Vol": 14, "Collagen Powder Vol": 14,
        "Lg Incision Likelihood": 16, "Medical School": 28, "HCP URL": 30,
        "Subject Line": 35, "Draft Email": 60,
        "Call 1 Date": 12, "Call 1 Outcome": 20, "Call 1 Notes": 30,
        "Call 2 Date": 12, "Call 2 Outcome": 20, "Call 2 Notes": 30,
        "Call 3 Date": 12, "Call 3 Outcome": 20, "Call 3 Notes": 30,
        "Email 1 Date": 12, "Email 1 Outcome": 20, "Email 1 Notes": 30,
        "Email 2 Date": 12, "Email 2 Outcome": 20, "Email 2 Notes": 30,
        "Next Action": 20, "Next Action Date": 14,
        "Decision Maker?": 14, "Notes": 40,
    }

    # Create tier tabs for Spine & Neuro
    logger.info("\nCreating Spine & Neuro tier tabs...")
    sn_by_tier = {}
    for row in sn_rows:
        tier = row.get("Tier", "Unknown")
        sn_by_tier.setdefault(tier, []).append(row)

    for tier in TIERS:
        tier_rows = sn_by_tier.get(tier, [])
        short = tier.replace(" (0-30 min)", "").replace(" (30-60 min)", "").replace(" (60-120 min)", "").replace(" (120-180 min)", "").replace(" (180+ drivable)", "").replace(" (Requires flight)", "")
        tab_name = f"S&N {short}"
        if tier == "Hospital-Based":
            tab_name = "S&N Hospital"
        write_tier_sheet(wb, tab_name, sn_headers, tier_rows, col_widths, "Open Spine Vol")

    # Create tier tabs for Outside Ortho & Spine
    logger.info("\nCreating Outside Ortho tier tabs...")
    oo_by_tier = {}
    for row in oo_rows:
        tier = row.get("Tier", "Unknown")
        oo_by_tier.setdefault(tier, []).append(row)

    for tier in TIERS:
        tier_rows = oo_by_tier.get(tier, [])
        short = tier.replace(" (0-30 min)", "").replace(" (30-60 min)", "").replace(" (60-120 min)", "").replace(" (120-180 min)", "").replace(" (180+ drivable)", "").replace(" (Requires flight)", "")
        tab_name = f"OOS {short}"
        if tier == "Hospital-Based":
            tab_name = "OOS Hospital"
        write_tier_sheet(wb, tab_name, oo_headers, tier_rows, col_widths, "Procedure Vol")

    # Save
    wb.save(input_file)
    logger.info(f"\nSaved to {input_file}")
    logger.info(f"Total tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
