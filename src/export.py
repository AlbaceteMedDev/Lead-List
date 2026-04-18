"""Excel workbook emitter — Master Lead List Tracker structure.

Sheet layout:
    Dashboard
    Top Targets - JR
    Top Targets - S&N
    Top Targets - OOS
    Call Tracker  - JR
    Email Tracker - JR
    Email Drafts  - JR
    Call Tracker  - S&N
    Email Tracker - S&N
    Email Drafts  - S&N
    Call Tracker  - OOS
    Email Tracker - OOS
    Email Drafts  - OOS
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterable

import pandas as pd

from src import routing

log = logging.getLogger(__name__)

CALL_TRACKER_COLS = [
    "HCP NPI", "First Name", "Last Name", "Credential", "Specialty",
    "Verified Phone", "Phone Status",
    "Primary Site of Care", "City", "State", "Practice Type", "Tier",
    "Lead Priority", "Lead Status", "MAC Jurisdiction", "Microlyte Eligible",
    "Joint Repl Vol", "Open Spine Vol", "Open Ortho Vol",
    "Lg Collagen Vol", "Sm/Md Collagen Vol", "Collagen Powder Vol",
    "Lg Incision Likelihood",
    "Call 1 Date", "Call 1 Outcome", "Call 1 Notes",
    "Call 2 Date", "Call 2 Outcome", "Call 2 Notes",
    "Call 3 Date", "Call 3 Outcome", "Call 3 Notes",
    "Call 4 Date", "Call 4 Outcome", "Call 4 Notes",
    "Call 5 Date", "Call 5 Outcome", "Call 5 Notes",
    "Next Action", "Next Action Date", "Decision Maker?",
    "Target Score", "Target Tier", "Why Target?", "Best Approach",
]

EMAIL_TRACKER_COLS = [
    "HCP NPI", "First Name", "Last Name", "Credential", "Specialty",
    "Email", "Email Status",
    "Primary Site of Care", "City", "State", "Practice Type", "Tier",
    "Lead Priority", "Lead Status", "MAC Jurisdiction", "Microlyte Eligible",
    "Joint Repl Vol", "Open Spine Vol", "Open Ortho Vol",
    "Email 1 Date", "Email 1 Subject", "Email 1 Outcome", "Email 1 Notes",
    "Email 2 Date", "Email 2 Subject", "Email 2 Outcome", "Email 2 Notes",
    "Email 3 Date", "Email 3 Subject", "Email 3 Outcome", "Email 3 Notes",
    "Next Action", "Next Action Date",
]

EMAIL_DRAFTS_COLS = ["HCP NPI", "Last Name", "Subject Line", "Draft Email"]

TOP_TARGETS_COLS = CALL_TRACKER_COLS[:]

COLUMN_WIDTHS = {
    "HCP NPI": 12, "First Name": 12, "Last Name": 14, "Credential": 10,
    "Specialty": 28, "Email": 32, "Email Status": 24,
    "Verified Phone": 13, "Phone Status": 20,
    "Primary Site of Care": 28, "City": 14, "State": 6,
    "Practice Type": 16, "Tier": 22, "Lead Priority": 10, "Lead Status": 18,
    "MAC Jurisdiction": 12, "Microlyte Eligible": 10,
    "Joint Repl Vol": 11, "Open Spine Vol": 11, "Open Ortho Vol": 11,
    "Lg Collagen Vol": 11, "Sm/Md Collagen Vol": 13, "Collagen Powder Vol": 13,
    "Lg Incision Likelihood": 13,
    "Next Action": 24, "Next Action Date": 13, "Decision Maker?": 10,
    "Target Score": 10, "Target Tier": 9, "Why Target?": 40, "Best Approach": 22,
    "Subject Line": 36, "Draft Email": 60,
}
for i in range(1, 6):
    COLUMN_WIDTHS[f"Call {i} Date"] = 12
    COLUMN_WIDTHS[f"Call {i} Outcome"] = 18
    COLUMN_WIDTHS[f"Call {i} Notes"] = 30
for i in range(1, 4):
    COLUMN_WIDTHS[f"Email {i} Date"] = 12
    COLUMN_WIDTHS[f"Email {i} Subject"] = 28
    COLUMN_WIDTHS[f"Email {i} Outcome"] = 18
    COLUMN_WIDTHS[f"Email {i} Notes"] = 30


CALL_OUTCOMES = [
    "No Answer", "Voicemail", "Gatekeeper - Declined", "Gatekeeper - Gave Info",
    "Wrong Number", "Bad Number", "Do Not Call", "Connected - Not Interested",
    "Connected - Interested", "Meeting Booked", "Callback Requested",
]
EMAIL_OUTCOMES = [
    "Sent", "Bounced", "Opened", "Replied - Interested",
    "Replied - Not Interested", "Meeting Booked", "Unsubscribed",
]
LEAD_STATUSES = [
    "New", "Queued", "Attempting Contact", "Connected", "Interested",
    "Meeting Booked", "Nurture", "Not Interested", "Do Not Contact",
    "Closed - Won", "Closed - Lost",
]
DECISION_MAKER_OPTIONS = ["Yes", "No", "Unknown"]


def _ensure_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    df = df.copy()
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df[columns]


def _sort_by_score(df: pd.DataFrame) -> pd.DataFrame:
    if "Target Score" in df.columns:
        scored = pd.to_numeric(df["Target Score"], errors="coerce").fillna(0)
        return df.assign(_score=scored).sort_values("_score", ascending=False).drop(columns="_score")
    return df


def _dashboard_metrics(df: pd.DataFrame, by_line: dict[str, pd.DataFrame]) -> list[list]:
    def count(subset: pd.DataFrame, mask: pd.Series | None = None) -> int:
        if subset.empty:
            return 0
        return int(mask.sum()) if mask is not None else len(subset)

    def by_status(subset: pd.DataFrame, status: str) -> int:
        return count(subset, subset.get("Lead Status", pd.Series(dtype=str)).fillna("") == status)

    def total_calls(subset: pd.DataFrame) -> int:
        n = 0
        for i in range(1, 6):
            col = f"Call {i} Date"
            if col in subset.columns:
                n += int((subset[col].astype(str).str.strip() != "").sum())
        return n

    def total_emails(subset: pd.DataFrame) -> int:
        n = 0
        for i in range(1, 4):
            col = f"Email {i} Date"
            if col in subset.columns:
                n += int((subset[col].astype(str).str.strip() != "").sum())
        return n

    lines = [code for code in routing.PRODUCT_LINES if code in by_line]
    rows: list[list] = [
        ["ALBACETE MEDDEV — OUTREACH DASHBOARD"],
        ["Cold Outreach Performance", "Regenerated on pipeline run"],
        [],
        ["OVERVIEW"],
        ["Total Leads", count(df), "Private Practices", count(df, df.get("Practice Type", pd.Series()) == "Private Practice")],
        ["Microlyte Eligible", count(df, df.get("Microlyte Eligible", pd.Series()) == "Yes"),
         "LCD-Blocked", count(df, df.get("Microlyte Eligible", pd.Series()) == "No")],
        [],
        ["ACTIVITY — ALL PORTFOLIOS"],
        ["Total Calls Logged", total_calls(df), "Total Emails Logged", total_emails(df)],
        ["Meetings Booked", by_status(df, "Meeting Booked"), "Interested", by_status(df, "Interested")],
        ["Connected", by_status(df, "Connected"), "Not Interested", by_status(df, "Not Interested")],
        ["Do Not Contact", by_status(df, "Do Not Contact"), "Nurture", by_status(df, "Nurture")],
        [],
        ["SIDE-BY-SIDE — JR vs S&N vs OOS"],
        ["Metric"] + [f"{code} ({routing.PRODUCT_LINE_LABELS[code]})" for code in lines],
        ["Total Leads"] + [count(by_line[c]) for c in lines],
        ["Private Practice"] + [count(by_line[c], by_line[c].get("Practice Type", pd.Series()) == "Private Practice") for c in lines],
        ["Hospital-Based"] + [count(by_line[c], by_line[c].get("Practice Type", pd.Series()) == "Hospital-Based") for c in lines],
        ["Microlyte Eligible"] + [count(by_line[c], by_line[c].get("Microlyte Eligible", pd.Series()) == "Yes") for c in lines],
        ["Tier 1 (0-30 min)"] + [count(by_line[c], by_line[c].get("Tier", pd.Series()) == "Tier 1 (0-30 min)") for c in lines],
        ["Tier 2 (30-60 min)"] + [count(by_line[c], by_line[c].get("Tier", pd.Series()) == "Tier 2 (30-60 min)") for c in lines],
        ["Calls Logged"] + [total_calls(by_line[c]) for c in lines],
        ["Emails Logged"] + [total_emails(by_line[c]) for c in lines],
        ["Meetings Booked"] + [by_status(by_line[c], "Meeting Booked") for c in lines],
        ["Target Tier A+"] + [count(by_line[c], by_line[c].get("Target Tier", pd.Series()) == "A+") for c in lines],
        ["Target Tier A"] + [count(by_line[c], by_line[c].get("Target Tier", pd.Series()) == "A") for c in lines],
        [],
        ["INCISION LIKELIHOOD"],
        ["Metric"] + [f"{code}" for code in lines],
        ["High"] + [count(by_line[c], by_line[c].get("Lg Incision Likelihood", pd.Series()) == "High") for c in lines],
        ["Medium-High"] + [count(by_line[c], by_line[c].get("Lg Incision Likelihood", pd.Series()) == "Medium-High") for c in lines],
        ["Medium"] + [count(by_line[c], by_line[c].get("Lg Incision Likelihood", pd.Series()) == "Medium") for c in lines],
        ["Low"] + [count(by_line[c], by_line[c].get("Lg Incision Likelihood", pd.Series()) == "Low") for c in lines],
    ]
    return rows


def _write_sheet(
    ws, columns: list[str], frame: pd.DataFrame, styles: dict,
    data_validations: dict[str, list[str]] | None = None,
) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_font, header_fill, header_align = styles["header_font"], styles["header_fill"], styles["header_align"]
    data_font, alt_fill, border = styles["data_font"], styles["alt_fill"], styles["border"]

    for idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(name, 15)

    for r_idx, (_, row) in enumerate(frame.iterrows(), start=2):
        for c_idx, name in enumerate(columns, start=1):
            value = row[name]
            if isinstance(value, str) and value == "":
                value = None
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = data_font
            cell.border = border
            if r_idx % 2 == 0:
                cell.fill = alt_fill
            if name in ("Why Target?", "Draft Email", "Call 1 Notes", "Call 2 Notes", "Call 3 Notes",
                        "Call 4 Notes", "Call 5 Notes", "Email 1 Notes", "Email 2 Notes", "Email 3 Notes",
                        "Next Action"):
                cell.alignment = styles["wrap_top"]

    ws.freeze_panes = "E2"
    if ws.max_row >= 2:
        ws.auto_filter.ref = ws.dimensions

    if data_validations:
        for col_name, options in data_validations.items():
            if col_name not in columns:
                continue
            col_idx = columns.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            formula = '"' + ",".join(options) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.add(f"{col_letter}2:{col_letter}{max(ws.max_row, 2)}")
            ws.add_data_validation(dv)


def _build_styles():
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    thin = Side(border_style="thin", color="FFD5D8DC")
    return {
        "header_font": Font(name="Arial", size=10, bold=True, color="FFFFFFFF"),
        "header_fill": PatternFill("solid", fgColor="FF1B4F72"),
        "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "data_font": Font(name="Arial", size=9),
        "alt_fill": PatternFill("solid", fgColor="FFEBF5FB"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "wrap_top": Alignment(wrap_text=True, vertical="top"),
    }


def _write_dashboard(ws, df: pd.DataFrame, by_line: dict[str, pd.DataFrame], styles: dict) -> None:
    from openpyxl.styles import Font

    rows = _dashboard_metrics(df, by_line)
    title_font = Font(name="Arial", size=14, bold=True, color="FF1B4F72")
    section_font = Font(name="Arial", size=11, bold=True, color="FF1B4F72")
    data_font = styles["data_font"]

    ws.column_dimensions["A"].width = 34
    for letter in ("B", "C", "D"):
        ws.column_dimensions[letter].width = 22

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = data_font
        first = row[0] if row else ""
        if isinstance(first, str):
            if first == "ALBACETE MEDDEV — OUTREACH DASHBOARD":
                ws.cell(row=r_idx, column=1).font = title_font
            elif first.isupper() and len(first) > 3 and r_idx > 1:
                ws.cell(row=r_idx, column=1).font = section_font
    ws.freeze_panes = "A2"


def write_workbook(df: pd.DataFrame, output_path: Path, top_targets_limit: int = 250) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    styles = _build_styles()

    df = df.copy()
    if "Product Line" not in df.columns:
        df = routing.enrich_frame(df)
    by_line = routing.split_by_product_line(df)

    validations = {
        "Lead Status": LEAD_STATUSES,
        "Decision Maker?": DECISION_MAKER_OPTIONS,
    }
    call_validations = dict(validations)
    for i in range(1, 6):
        call_validations[f"Call {i} Outcome"] = CALL_OUTCOMES
    email_validations = dict(validations)
    for i in range(1, 4):
        email_validations[f"Email {i} Outcome"] = EMAIL_OUTCOMES

    dashboard_ws = wb.create_sheet("Dashboard")
    _write_dashboard(dashboard_ws, df, by_line, styles)

    for code in routing.PRODUCT_LINES:
        if code not in by_line:
            continue
        subset = by_line[code]
        top = _sort_by_score(subset).head(top_targets_limit)
        ws = wb.create_sheet(f"Top Targets - {code}")
        _write_sheet(ws, TOP_TARGETS_COLS, _ensure_columns(top, TOP_TARGETS_COLS), styles, call_validations)

    for code in routing.PRODUCT_LINES:
        if code not in by_line:
            continue
        subset = _sort_by_score(by_line[code])

        call_ws = wb.create_sheet(f"Call Tracker - {code}")
        _write_sheet(call_ws, CALL_TRACKER_COLS, _ensure_columns(subset, CALL_TRACKER_COLS), styles, call_validations)

        email_ws = wb.create_sheet(f"Email Tracker - {code}")
        _write_sheet(email_ws, EMAIL_TRACKER_COLS, _ensure_columns(subset, EMAIL_TRACKER_COLS), styles, email_validations)

        drafts_ws = wb.create_sheet(f"Email Drafts - {code}")
        _write_sheet(drafts_ws, EMAIL_DRAFTS_COLS, _ensure_columns(subset, EMAIL_DRAFTS_COLS), styles)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Wrote workbook: %s", output_path)
    return output_path
