"""Rebuild Dashboard with a clean, easy-to-scan KPI card layout."""

import logging
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("dashboard")

# ============================================================
# STYLE SYSTEM
# ============================================================

# Colors
NAVY = "1F4E78"
PURPLE = "7030A0"
RED = "C00000"
TEAL = "006666"
DARK_GRAY = "3B3B3B"
LIGHT_GRAY_BG = "F2F2F2"
MED_GRAY_BG = "E7E6E6"
WHITE = "FFFFFF"

# Portfolio colors
JR_COLOR = NAVY
SN_COLOR = PURPLE
OOS_COLOR = RED

# Fonts
TITLE_FONT = Font(name="Calibri", size=22, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="808080")
SECTION_BANNER_FONT = Font(name="Calibri", size=14, bold=True, color=WHITE)
SUBSECTION_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_GRAY)

KPI_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="595959")
KPI_VALUE_FONT_LG = Font(name="Calibri", size=20, bold=True, color=DARK_GRAY)
KPI_VALUE_FONT_MD = Font(name="Calibri", size=14, bold=True, color=DARK_GRAY)
KPI_VALUE_FONT_ACCENT_JR = Font(name="Calibri", size=14, bold=True, color=JR_COLOR)
KPI_VALUE_FONT_ACCENT_SN = Font(name="Calibri", size=14, bold=True, color=SN_COLOR)
KPI_VALUE_FONT_ACCENT_OOS = Font(name="Calibri", size=14, bold=True, color=OOS_COLOR)

TABLE_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
TABLE_CELL_FONT = Font(name="Calibri", size=11, bold=True, color=DARK_GRAY)

# Fills
TITLE_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
JR_BANNER_FILL = PatternFill(start_color=JR_COLOR, end_color=JR_COLOR, fill_type="solid")
SN_BANNER_FILL = PatternFill(start_color=SN_COLOR, end_color=SN_COLOR, fill_type="solid")
OOS_BANNER_FILL = PatternFill(start_color=OOS_COLOR, end_color=OOS_COLOR, fill_type="solid")
TEAL_BANNER_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")

CARD_FILL = PatternFill(start_color=LIGHT_GRAY_BG, end_color=LIGHT_GRAY_BG, fill_type="solid")
TABLE_HEADER_FILL = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
TABLE_ROW_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
WHITE_FILL_PATT = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
TABLE_ALT_FILL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

# Borders
THIN_LIGHT = Side(style="thin", color="D0D0D0")
THICK_DARK = Side(style="medium", color=DARK_GRAY)
CARD_BORDER = Border(left=THIN_LIGHT, right=THIN_LIGHT, top=THIN_LIGHT, bottom=THIN_LIGHT)
TABLE_BORDER = Border(left=THIN_LIGHT, right=THIN_LIGHT, top=THIN_LIGHT, bottom=THIN_LIGHT)

# Alignment
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_INDENT = Alignment(horizontal="left", vertical="center", indent=1)


# ============================================================
# BUILDING BLOCKS
# ============================================================

def draw_title(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 42


def draw_subtitle(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18


def draw_section_banner(ws, row, text, fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_BANNER_FONT
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28


def draw_subsection_header(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBSECTION_FONT
    cell.alignment = LEFT_INDENT
    ws.row_dimensions[row].height = 20


def draw_kpi_card(ws, row, col_start, col_end, label, formula, value_font=None):
    """Draw a KPI card: label on top, value on bottom."""
    if value_font is None:
        value_font = KPI_VALUE_FONT_LG

    # Label row
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    label_cell = ws.cell(row=row, column=col_start, value=label)
    label_cell.font = KPI_LABEL_FONT
    label_cell.fill = CARD_FILL
    label_cell.alignment = CENTER
    label_cell.border = CARD_BORDER

    # Apply border to all merged cells
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = CARD_BORDER
        ws.cell(row=row, column=c).fill = CARD_FILL

    # Value row
    ws.merge_cells(start_row=row + 1, start_column=col_start, end_row=row + 1, end_column=col_end)
    val_cell = ws.cell(row=row + 1, column=col_start, value=formula)
    val_cell.font = value_font
    val_cell.fill = CARD_FILL
    val_cell.alignment = CENTER
    val_cell.border = CARD_BORDER
    val_cell.number_format = "#,##0"
    for c in range(col_start, col_end + 1):
        ws.cell(row=row + 1, column=c).border = CARD_BORDER
        ws.cell(row=row + 1, column=c).fill = CARD_FILL

    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 34


def draw_kpi_card_pct(ws, row, col_start, col_end, label, formula, value_font=None):
    """KPI card with percentage formatting."""
    if value_font is None:
        value_font = KPI_VALUE_FONT_LG
    draw_kpi_card(ws, row, col_start, col_end, label, formula, value_font)
    val_cell = ws.cell(row=row + 1, column=col_start)
    val_cell.number_format = "0.0%"


def draw_mini_table(ws, start_row, start_col, headers, values, widths=None):
    """Draw a clean mini-table with header row on top."""
    # Header row
    for i, h in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=h)
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = CENTER
        cell.border = TABLE_BORDER
    ws.row_dimensions[start_row].height = 28

    # Value row
    for i, v in enumerate(values):
        cell = ws.cell(row=start_row + 1, column=start_col + i, value=v)
        cell.font = TABLE_CELL_FONT
        cell.fill = TABLE_ROW_FILL
        cell.alignment = CENTER
        cell.border = TABLE_BORDER
        cell.number_format = "#,##0"
    ws.row_dimensions[start_row + 1].height = 26


def draw_mini_table_pct(ws, start_row, start_col, headers, values, pct_indices):
    """Mini-table with some columns formatted as percentage."""
    draw_mini_table(ws, start_row, start_col, headers, values)
    for pi in pct_indices:
        cell = ws.cell(row=start_row + 1, column=start_col + pi)
        cell.number_format = "0.0%"


# ============================================================
# PORTFOLIO SECTION BUILDER
# ============================================================

def build_portfolio_section(ws, start_row, portfolio_name, color_fill,
                            call_tab, email_tab, total_range, vol_col, vol_label,
                            accent_font, collagen_stats=None):
    """Build a complete portfolio section: banner + KPIs + mini-tables."""
    row = start_row

    # Section banner
    draw_section_banner(ws, row, f"  ▍ {portfolio_name}", color_fill)
    row += 2

    # KPI cards row (4 cards, 3 cols each)
    t = call_tab
    tr = total_range

    draw_kpi_card(ws, row, 1, 3, "TOTAL LEADS",
                  f"=COUNTA('{t}'!A2:A{tr})", accent_font)
    draw_kpi_card(ws, row, 4, 6, "PRIVATE PRACTICE",
                  f"=COUNTIF('{t}'!K2:K{tr},\"Private*\")", accent_font)
    draw_kpi_card(ws, row, 7, 9, "HOSPITAL-BASED",
                  f"=COUNTIF('{t}'!K2:K{tr},\"Hospital*\")", accent_font)
    draw_kpi_card(ws, row, 10, 12, "MICROLYTE ELIGIBLE",
                  f"=COUNTIF('{t}'!P2:P{tr},\"Yes\")", accent_font)
    row += 3

    # Second KPI row — drive time tiers
    draw_kpi_card(ws, row, 1, 3, "TIER 1 (0-30 MIN)",
                  f"=COUNTIF('{t}'!L2:L{tr},\"Tier 1*\")", KPI_VALUE_FONT_MD)
    draw_kpi_card(ws, row, 4, 6, "TIER 2 (30-60 MIN)",
                  f"=COUNTIF('{t}'!L2:L{tr},\"Tier 2*\")", KPI_VALUE_FONT_MD)
    draw_kpi_card(ws, row, 7, 9, "TIER 3-4 (1-3 HRS)",
                  f"=COUNTIF('{t}'!L2:L{tr},\"Tier 3*\")+COUNTIF('{t}'!L2:L{tr},\"Tier 4*\")",
                  KPI_VALUE_FONT_MD)
    draw_kpi_card(ws, row, 10, 12, vol_label.upper(),
                  f"=IFERROR(ROUND(AVERAGE('{t}'!{vol_col}2:{vol_col}{tr}),0),0)",
                  KPI_VALUE_FONT_MD)
    row += 3

    # ---- PIPELINE STATUS ----
    draw_subsection_header(ws, row, "  Pipeline Status")
    row += 1
    draw_mini_table(ws, row, 1,
                    ["New", "Contacted", "Mtg Sched", "Mtg Done", "Proposal", "Won", "Lost", "Nurture"],
                    [
                        f"=COUNTIF('{t}'!N2:N{tr},\"New\")",
                        f"=COUNTIF('{t}'!N2:N{tr},\"Contacted\")",
                        f"=COUNTIF('{t}'!N2:N{tr},\"Meeting Scheduled\")",
                        f"=COUNTIF('{t}'!N2:N{tr},\"Meeting Completed\")",
                        f"=COUNTIF('{t}'!N2:N{tr},\"Proposal Sent\")",
                        f"=COUNTIF('{t}'!N2:N{tr},\"Won\")",
                        f"=COUNTIF('{t}'!N2:N{tr},\"Lost\")",
                        f"=COUNTIF('{t}'!N2:N{tr},\"Nurture\")",
                    ])
    row += 3

    # ---- CALL ACTIVITY ----
    draw_subsection_header(ws, row, "  Call Activity")
    row += 1
    outcome_cols = ["W", "Z", "AC", "AF", "AI"]
    total_calls_f = "=" + "+".join(f"COUNTA('{t}'!{c}2:{c}{tr})" for c in outcome_cols)
    connected_f = "=" + "+".join(f"COUNTIF('{t}'!{c}2:{c}{tr},\"Connected*\")" for c in outcome_cols)
    voicemail_f = "=" + "+".join(f"COUNTIF('{t}'!{c}2:{c}{tr},\"Voicemail\")" for c in outcome_cols)
    no_answer_f = "=" + "+".join(f"COUNTIF('{t}'!{c}2:{c}{tr},\"No Answer\")" for c in outcome_cols)
    mtg_set_f = "=" + "+".join(f"COUNTIF('{t}'!{c}2:{c}{tr},\"Connected - Meeting Set\")" for c in outcome_cols)
    connect_rate_f = f"=IFERROR(({connected_f[1:]})/({total_calls_f[1:]}),0)"
    meeting_rate_f = f"=IFERROR(({mtg_set_f[1:]})/({total_calls_f[1:]}),0)"
    draw_mini_table_pct(ws, row, 1,
                        ["Total Calls", "Connected", "Voicemail", "No Answer",
                         "Meetings Set", "Connect Rate", "Meeting Rate", ""],
                        [total_calls_f, connected_f, voicemail_f, no_answer_f,
                         mtg_set_f, connect_rate_f, meeting_rate_f, ""],
                        pct_indices=[5, 6])
    row += 3

    # ---- EMAIL ACTIVITY ----
    draw_subsection_header(ws, row, "  Email Activity")
    row += 1
    et = email_tab
    email_outcome_cols = ["T", "X", "AB"]
    total_emails_f = "=" + "+".join(f"COUNTA('{et}'!{c}2:{c}{tr})" for c in email_outcome_cols)
    opened_f = "=" + "+".join(f"COUNTIF('{et}'!{c}2:{c}{tr},\"Opened\")" for c in email_outcome_cols)
    replied_f = "=" + "+".join(f"COUNTIF('{et}'!{c}2:{c}{tr},\"Replied\")" for c in email_outcome_cols)
    bounced_f = "=" + "+".join(f"COUNTIF('{et}'!{c}2:{c}{tr},\"Bounced\")" for c in email_outcome_cols)
    open_rate_f = f"=IFERROR(({opened_f[1:]})/({total_emails_f[1:]}),0)"
    reply_rate_f = f"=IFERROR(({replied_f[1:]})/({total_emails_f[1:]}),0)"
    draw_mini_table_pct(ws, row, 1,
                        ["Total Sent", "Opened", "Replied", "Bounced",
                         "Open Rate", "Reply Rate", "", ""],
                        [total_emails_f, opened_f, replied_f, bounced_f,
                         open_rate_f, reply_rate_f, "", ""],
                        pct_indices=[4, 5])
    row += 3

    # ---- COLLAGEN USAGE (uses cached values since it's static AcuityMD data) ----
    draw_subsection_header(ws, row, "  🧬 Collagen Usage (Existing Wound-Care Product Buyers)")
    row += 1
    cs = collagen_stats or {}
    # Big summary cards — use direct cached values so they show immediately
    draw_kpi_card(ws, row, 1, 3, "LEADS USING COLLAGEN (ANY TYPE)",
                  cs.get("any_collagen_users", 0), KPI_VALUE_FONT_MD)
    draw_kpi_card(ws, row, 4, 6, "TOTAL LG SHEET VOL",
                  cs.get("lg_total", 0), KPI_VALUE_FONT_MD)
    draw_kpi_card(ws, row, 7, 9, "TOTAL SM/MD SHEET VOL",
                  cs.get("smmd_total", 0), KPI_VALUE_FONT_MD)
    draw_kpi_card(ws, row, 10, 12, "TOTAL COLLAGEN POWDER VOL",
                  cs.get("pwd_total", 0), KPI_VALUE_FONT_MD)
    row += 3
    draw_mini_table(ws, row, 1,
                    ["Lg Sheet Users", "Sm/Md Sheet Users", "Powder Users",
                     "Avg Lg Sheet", "Avg Sm/Md Sheet", "Avg Powder",
                     "Max Lg Sheet", "Max Sm/Md"],
                    [
                        cs.get("lg_users", 0),
                        cs.get("smmd_users", 0),
                        cs.get("pwd_users", 0),
                        cs.get("lg_avg", 0),
                        cs.get("smmd_avg", 0),
                        cs.get("pwd_avg", 0),
                        cs.get("lg_max", 0),
                        cs.get("smmd_max", 0),
                    ])
    row += 3

    # ---- INCISION LIKELIHOOD ----
    draw_subsection_header(ws, row, "  Large Incision Likelihood")
    row += 1
    high_f = f"=COUNTIF('{t}'!U2:U{tr},\"High\")"
    medhigh_f = f"=COUNTIF('{t}'!U2:U{tr},\"Medium-High\")"
    medium_f = f"=COUNTIF('{t}'!U2:U{tr},\"Medium\")"
    low_f = f"=COUNTIF('{t}'!U2:U{tr},\"Low\")"
    unlikely_f = f"=COUNTIF('{t}'!U2:U{tr},\"Unlikely\")"
    draw_mini_table_pct(ws, row, 1,
                        ["High 🟢", "Medium-High", "Medium", "Low", "Unlikely",
                         "High+MH", "% High+MH", ""],
                        [high_f, medhigh_f, medium_f, low_f, unlikely_f,
                         f"=({high_f[1:]})+({medhigh_f[1:]})",
                         f"=IFERROR((({high_f[1:]})+({medhigh_f[1:]}))/COUNTA('{t}'!U2:U{tr}),0)",
                         ""],
                        pct_indices=[6])
    row += 4

    return row


# ============================================================
# MAIN
# ============================================================

def compute_stats(input_file):
    """Pre-compute statistics from each tracker tab for caching in Dashboard."""
    import pandas as pd
    import openpyxl as ox

    def safe_int(v):
        if v is None or pd.isna(v):
            return 0
        try:
            return int(v)
        except (ValueError, TypeError):
            return 0

    wb = ox.load_workbook(input_file, read_only=True)
    stats = {}
    for tab in ["Call Tracker - JR", "Call Tracker - S&N", "Call Tracker - OOS"]:
        ws = wb[tab]
        headers = [c.value for c in list(ws.iter_rows(max_row=1))[0]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        df = pd.DataFrame(rows, columns=headers)
        for col in ["Lg Collagen Vol", "Sm/Md Collagen Vol", "Collagen Powder Vol"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        lg_pos = df[df["Lg Collagen Vol"] > 0]["Lg Collagen Vol"]
        smmd_pos = df[df["Sm/Md Collagen Vol"] > 0]["Sm/Md Collagen Vol"]
        pwd_pos = df[df["Collagen Powder Vol"] > 0]["Collagen Powder Vol"]
        stats[tab] = {
            "lg_users": int((df["Lg Collagen Vol"] > 0).sum()),
            "smmd_users": int((df["Sm/Md Collagen Vol"] > 0).sum()),
            "pwd_users": int((df["Collagen Powder Vol"] > 0).sum()),
            "lg_total": safe_int(df["Lg Collagen Vol"].sum()),
            "smmd_total": safe_int(df["Sm/Md Collagen Vol"].sum()),
            "pwd_total": safe_int(df["Collagen Powder Vol"].sum()),
            "lg_avg": safe_int(lg_pos.mean() if len(lg_pos) else 0),
            "smmd_avg": safe_int(smmd_pos.mean() if len(smmd_pos) else 0),
            "pwd_avg": safe_int(pwd_pos.mean() if len(pwd_pos) else 0),
            "lg_max": safe_int(df["Lg Collagen Vol"].max()),
            "smmd_max": safe_int(df["Sm/Md Collagen Vol"].max()),
            "any_collagen_users": int((
                (df["Lg Collagen Vol"] > 0) |
                (df["Sm/Md Collagen Vol"] > 0) |
                (df["Collagen Powder Vol"] > 0)
            ).sum()),
        }
        logger.info(f"  {tab} collagen: {stats[tab]['any_collagen_users']} users, "
                    f"Lg total={stats[tab]['lg_total']}, Sm/Md total={stats[tab]['smmd_total']}, "
                    f"Pwd total={stats[tab]['pwd_total']}")
    wb.close()
    return stats


def main():
    input_file = "Master_Lead_List_Tracker (3).xlsx"

    logger.info("Pre-computing stats for cached values...")
    stats = compute_stats(input_file)

    logger.info("Loading workbook...")
    wb = openpyxl.load_workbook(input_file)

    # Strip external links
    if hasattr(wb, "_external_links"):
        wb._external_links = []

    # Force Excel to fully recalculate all formulas when the file opens
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    # Remove old Dashboard
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_properties.tabColor = NAVY

    # Set column widths — 12 cols, each relatively wide for readability
    col_widths = [15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hide gridlines
    ws.sheet_view.showGridLines = False

    # === TITLE ===
    row = 1
    draw_title(ws, row, "ALBACETE MEDDEV — OUTREACH DASHBOARD")
    row += 1
    draw_subtitle(ws, row, "Cold Outreach Performance | Updates automatically on open")
    row += 2

    # === TOP-LEVEL OVERVIEW ===
    draw_section_banner(ws, row, "  📊 OVERVIEW — ALL PORTFOLIOS COMBINED", TEAL_BANNER_FILL)
    row += 2

    # Big combined KPI cards (4 cards, 3 cols each)
    total_leads_f = ("=COUNTA('Call Tracker - JR'!A2:A4442)+"
                    "COUNTA('Call Tracker - S&N'!A2:A10001)+"
                    "COUNTA('Call Tracker - OOS'!A2:A10001)")
    total_pp_f = ("=COUNTIF('Call Tracker - JR'!K2:K4442,\"Private*\")+"
                 "COUNTIF('Call Tracker - S&N'!K2:K10001,\"Private*\")+"
                 "COUNTIF('Call Tracker - OOS'!K2:K10001,\"Private*\")")
    total_hospital_f = ("=COUNTIF('Call Tracker - JR'!K2:K4442,\"Hospital*\")+"
                       "COUNTIF('Call Tracker - S&N'!K2:K10001,\"Hospital*\")+"
                       "COUNTIF('Call Tracker - OOS'!K2:K10001,\"Hospital*\")")
    total_elig_f = ("=COUNTIF('Call Tracker - JR'!P2:P4442,\"Yes\")+"
                   "COUNTIF('Call Tracker - S&N'!P2:P10001,\"Yes\")+"
                   "COUNTIF('Call Tracker - OOS'!P2:P10001,\"Yes\")")

    draw_kpi_card(ws, row, 1, 3, "TOTAL LEADS", total_leads_f, KPI_VALUE_FONT_LG)
    draw_kpi_card(ws, row, 4, 6, "PRIVATE PRACTICES", total_pp_f, KPI_VALUE_FONT_LG)
    draw_kpi_card(ws, row, 7, 9, "HOSPITAL-BASED", total_hospital_f, KPI_VALUE_FONT_LG)
    draw_kpi_card(ws, row, 10, 12, "MICROLYTE ELIGIBLE", total_elig_f, KPI_VALUE_FONT_LG)
    row += 3

    # ===== SIDE-BY-SIDE PORTFOLIO COMPARISON TABLE =====
    draw_subsection_header(ws, row, "  📋 Side-by-Side Comparison — JR vs S&N vs OOS")
    row += 1

    # Column headers: metric + JR + S&N + OOS
    compare_headers = ["Metric", "JR (Joint Repl)", "S&N (Spine & Neuro)", "OOS (Outside Ortho)"]
    # We'll use cols 1-4 for this wider comparison; fill remaining cols with empty
    for i, h in enumerate(compare_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = CENTER
        cell.border = TABLE_BORDER
    # Merge the extra columns for a wider right-hand area (use cols 1,2,3,4 + optional notes)
    ws.row_dimensions[row].height = 26
    row += 1

    def compare_row(label, jr_val, sn_val, oos_val, is_pct=False):
        nonlocal row
        # Label
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.fill = TABLE_ALT_FILL
        cell.border = TABLE_BORDER
        # Values
        for col, v, color in [(2, jr_val, JR_COLOR), (3, sn_val, SN_COLOR), (4, oos_val, OOS_COLOR)]:
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(name="Calibri", size=11, bold=True, color=color)
            cell.alignment = CENTER
            cell.fill = TABLE_ROW_FILL
            cell.border = TABLE_BORDER
            cell.number_format = "0.0%" if is_pct else "#,##0"
        # Merge cols 5-12 as empty visual space
        for col in range(5, 13):
            ws.cell(row=row, column=col).border = None
            ws.cell(row=row, column=col).fill = WHITE_FILL_PATT
        ws.row_dimensions[row].height = 22
        row += 1

    # Widen cols 2-4 so comparison values breathe
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    jr_stats = stats.get("Call Tracker - JR", {})
    sn_stats = stats.get("Call Tracker - S&N", {})
    oos_stats = stats.get("Call Tracker - OOS", {})

    # Portfolio size
    compare_row("Total Leads",
                "=COUNTA('Call Tracker - JR'!A2:A4442)",
                "=COUNTA('Call Tracker - S&N'!A2:A10001)",
                "=COUNTA('Call Tracker - OOS'!A2:A10001)")
    compare_row("Private Practice",
                "=COUNTIF('Call Tracker - JR'!K2:K4442,\"Private*\")",
                "=COUNTIF('Call Tracker - S&N'!K2:K10001,\"Private*\")",
                "=COUNTIF('Call Tracker - OOS'!K2:K10001,\"Private*\")")
    compare_row("Hospital-Based",
                "=COUNTIF('Call Tracker - JR'!K2:K4442,\"Hospital*\")",
                "=COUNTIF('Call Tracker - S&N'!K2:K10001,\"Hospital*\")",
                "=COUNTIF('Call Tracker - OOS'!K2:K10001,\"Hospital*\")")
    compare_row("Microlyte Eligible (non-LCD states)",
                "=COUNTIF('Call Tracker - JR'!P2:P4442,\"Yes\")",
                "=COUNTIF('Call Tracker - S&N'!P2:P10001,\"Yes\")",
                "=COUNTIF('Call Tracker - OOS'!P2:P10001,\"Yes\")")
    compare_row("Tier 1 (0–30 min from NYC)",
                "=COUNTIF('Call Tracker - JR'!L2:L4442,\"Tier 1*\")",
                "=COUNTIF('Call Tracker - S&N'!L2:L10001,\"Tier 1*\")",
                "=COUNTIF('Call Tracker - OOS'!L2:L10001,\"Tier 1*\")")
    compare_row("Tier 2 (30–60 min)",
                "=COUNTIF('Call Tracker - JR'!L2:L4442,\"Tier 2*\")",
                "=COUNTIF('Call Tracker - S&N'!L2:L10001,\"Tier 2*\")",
                "=COUNTIF('Call Tracker - OOS'!L2:L10001,\"Tier 2*\")")

    # Divider row
    ws.row_dimensions[row].height = 8
    row += 1
    # Sub-header for collagen
    for i in range(1, 5):
        cell = ws.cell(row=row, column=i)
        cell.fill = PatternFill(start_color=MED_GRAY_BG, end_color=MED_GRAY_BG, fill_type="solid")
        cell.border = TABLE_BORDER
    ws.cell(row=row, column=1, value="  🧬 COLLAGEN USAGE").font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    compare_row("Leads Using Any Collagen",
                jr_stats.get("any_collagen_users", 0),
                sn_stats.get("any_collagen_users", 0),
                oos_stats.get("any_collagen_users", 0))
    compare_row("Total Lg Sheet Volume",
                jr_stats.get("lg_total", 0),
                sn_stats.get("lg_total", 0),
                oos_stats.get("lg_total", 0))
    compare_row("Total Sm/Md Sheet Volume",
                jr_stats.get("smmd_total", 0),
                sn_stats.get("smmd_total", 0),
                oos_stats.get("smmd_total", 0))
    compare_row("Total Powder Volume",
                jr_stats.get("pwd_total", 0),
                sn_stats.get("pwd_total", 0),
                oos_stats.get("pwd_total", 0))
    compare_row("Avg Lg Sheet (where > 0)",
                jr_stats.get("lg_avg", 0),
                sn_stats.get("lg_avg", 0),
                oos_stats.get("lg_avg", 0))
    compare_row("Max Lg Sheet Volume (top user)",
                jr_stats.get("lg_max", 0),
                sn_stats.get("lg_max", 0),
                oos_stats.get("lg_max", 0))

    # Divider row
    ws.row_dimensions[row].height = 8
    row += 1
    # Sub-header for incision
    for i in range(1, 5):
        cell = ws.cell(row=row, column=i)
        cell.fill = PatternFill(start_color=MED_GRAY_BG, end_color=MED_GRAY_BG, fill_type="solid")
        cell.border = TABLE_BORDER
    ws.cell(row=row, column=1, value="  🔪 INCISION LIKELIHOOD (Large Incision Targets)").font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    compare_row("High Incision Likelihood",
                "=COUNTIF('Call Tracker - JR'!U2:U4442,\"High\")",
                "=COUNTIF('Call Tracker - S&N'!U2:U10001,\"High\")",
                "=COUNTIF('Call Tracker - OOS'!U2:U10001,\"High\")")
    compare_row("Medium-High Incision",
                "=COUNTIF('Call Tracker - JR'!U2:U4442,\"Medium-High\")",
                "=COUNTIF('Call Tracker - S&N'!U2:U10001,\"Medium-High\")",
                "=COUNTIF('Call Tracker - OOS'!U2:U10001,\"Medium-High\")")
    compare_row("Medium",
                "=COUNTIF('Call Tracker - JR'!U2:U4442,\"Medium\")",
                "=COUNTIF('Call Tracker - S&N'!U2:U10001,\"Medium\")",
                "=COUNTIF('Call Tracker - OOS'!U2:U10001,\"Medium\")")
    compare_row("Low",
                "=COUNTIF('Call Tracker - JR'!U2:U4442,\"Low\")",
                "=COUNTIF('Call Tracker - S&N'!U2:U10001,\"Low\")",
                "=COUNTIF('Call Tracker - OOS'!U2:U10001,\"Low\")")
    compare_row("Unlikely",
                "=COUNTIF('Call Tracker - JR'!U2:U4442,\"Unlikely\")",
                "=COUNTIF('Call Tracker - S&N'!U2:U10001,\"Unlikely\")",
                "=COUNTIF('Call Tracker - OOS'!U2:U10001,\"Unlikely\")")

    row += 2

    # === JR PORTFOLIO ===
    row = build_portfolio_section(
        ws, row, "JOINT REPLACEMENT — JR",
        JR_BANNER_FILL, "Call Tracker - JR", "Email Tracker - JR",
        "4442", "Q", "Avg Joint Vol", KPI_VALUE_FONT_ACCENT_JR,
        collagen_stats=stats.get("Call Tracker - JR"),
    )

    # === S&N PORTFOLIO ===
    row += 1
    row = build_portfolio_section(
        ws, row, "SPINE & NEURO — S&N",
        SN_BANNER_FILL, "Call Tracker - S&N", "Email Tracker - S&N",
        "10001", "Q", "Avg Spine Vol", KPI_VALUE_FONT_ACCENT_SN,
        collagen_stats=stats.get("Call Tracker - S&N"),
    )

    # === OOS PORTFOLIO ===
    row += 1
    row = build_portfolio_section(
        ws, row, "OUTSIDE ORTHO & SPINE — OOS",
        OOS_BANNER_FILL, "Call Tracker - OOS", "Email Tracker - OOS",
        "10001", "Q", "Avg Proc Vol", KPI_VALUE_FONT_ACCENT_OOS,
        collagen_stats=stats.get("Call Tracker - OOS"),
    )

    # === READING GUIDE ===
    row += 1
    draw_section_banner(ws, row, "  ❓ HOW TO READ THIS DASHBOARD", TEAL_BANNER_FILL)
    row += 2
    guide_notes = [
        ("Total Leads", "Count of unique physicians in each portfolio — the pool you can call/email."),
        ("Private Practice vs Hospital-Based", "Private practices are your primary targets — easier to reach decision-makers."),
        ("Microlyte Eligible", "Leads in non-LCD states where you can pitch Microlyte SAM alongside ProPacks. LCD states get ProPacks-only emails."),
        ("Tier 1/2/3/4", "Drive-time tiers from NYC Midtown. Tier 1 = 0–30 min, Tier 2 = 30–60 min. Start with Tier 1-2 Private Practices for in-person visits."),
        ("Pipeline Status", "Updates automatically as you change the 'Lead Status' dropdown in each Call Tracker tab."),
        ("Call Activity", "Aggregated across all 5 call attempts per lead. Populated from 'Call 1-5 Outcome' dropdowns."),
        ("Email Activity", "Aggregated across all 3 email attempts. Populated from 'Email 1-3 Outcome' dropdowns."),
        ("Collagen Usage", "Procedure volume of current collagen sheet/powder products — signals existing wound-product buyers."),
        ("Incision Likelihood", "Predicted surgical incision size based on procedure mix. High/Medium-High = best Microlyte targets."),
    ]
    for label, desc in guide_notes:
        ws.cell(row=row, column=1, value=f"  {label}").font = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        desc_cell = ws.cell(row=row, column=2, value=desc)
        desc_cell.font = Font(name="Calibri", size=10, color="595959")
        desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22
        row += 1

    # Freeze panes so title stays visible
    ws.freeze_panes = "A4"

    wb.save(input_file)
    logger.info(f"Dashboard rebuilt in {input_file}")


if __name__ == "__main__":
    main()
