"""Enrich Spine & Neuro + Outside Ortho & Spine tabs, write back to tracker xlsx."""

import copy
import json
import logging
import os
import sys

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.classify import classify_practices
from src.tier import assign_tiers
from src.nppes import verify_phones
from src.email_enrich import enrich_emails
from src.mac_mapping import map_mac_jurisdictions

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("enrich_neuro")

# Styles
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Arial", size=9)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def extract_tab_to_df(wb, sheet_name):
    """Extract a worksheet into a pandas DataFrame."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = list(rows[0])
    data = [list(r) for r in rows[1:]]
    df = pd.DataFrame(data, columns=headers)
    # Ensure string types for key columns
    for col in ["HCP NPI", "Phone", "Email", "State", "Postal Code"]:
        if col in df.columns:
            df[col] = df[col].astype(str).replace({"None": None, "nan": None, "": None})
    return df


def enrich_dataframe(df, tab_name, force_nppes=False):
    """Run the full enrichment pipeline on a DataFrame."""
    # Normalize column names to match pipeline expectations
    col_map = {
        "Phone": "Phone Number",
        "Address": "Address 1",
    }
    df = df.rename(columns=col_map)

    logger.info(f"\n{'='*60}")
    logger.info(f"Enriching: {tab_name} ({len(df)} leads)")
    logger.info(f"{'='*60}")

    # Step 1: Practice Classification
    logger.info(">>> Classifying practices...")
    df = classify_practices(df)

    # Step 2: Drive-Time Tiering
    logger.info(">>> Assigning drive-time tiers...")
    df = assign_tiers(df)

    # Step 3: NPPES Phone Verification
    logger.info(">>> NPPES phone verification...")
    df = verify_phones(df, force=force_nppes)

    # Step 4: Email Enrichment
    logger.info(">>> Email enrichment...")
    df = enrich_emails(df)

    # Step 5: MAC Jurisdiction Mapping
    logger.info(">>> MAC jurisdiction mapping...")
    df = map_mac_jurisdictions(df)

    # Rename back
    reverse_map = {v: k for k, v in col_map.items()}
    df = df.rename(columns=reverse_map)

    return df


def compute_priority(row, p25, p75):
    """Compute lead priority A/B/C."""
    ptype = str(row.get("Practice Type", ""))
    tier = str(row.get("Tier", ""))
    vol = row.get("_total_vol", 0)

    is_private = ptype == "Private Practice"
    is_tier_12 = "Tier 1" in tier or "Tier 2" in tier
    is_tier_34 = "Tier 3" in tier or "Tier 4" in tier

    if is_private and is_tier_12 and vol >= p75:
        return "A"
    if is_private and is_tier_34:
        return "B"
    if is_private and is_tier_12 and vol >= p25:
        return "B"
    if vol >= p25 and vol < p75:
        return "B"
    return "C"


def generate_spine_neuro_email(row, templates_config):
    """Generate outreach email for Spine & Neuro leads."""
    sender = templates_config["sender"]
    last = row.get("Last Name", "")
    first = row.get("First Name", "")
    practice = row.get("Primary Site of Care") or "your practice"
    city = row.get("City") or ""
    specialty = str(row.get("Specialty") or "")
    microlyte = str(row.get("Microlyte Eligible", "No"))

    vol = _safe_float(row.get("Open Spine Vol"))
    if vol > 500:
        hook = f"With the volume of spine cases you're managing at {practice}"
    elif vol >= 100:
        hook = f"Given your spine surgery practice at {practice}"
    else:
        hook = f"As a surgeon performing spine procedures at {practice}"

    if "Neuro" in specialty:
        spec_ref = "neurosurgeon"
    elif "Spine" in specialty:
        spec_ref = "spine surgeon"
    else:
        spec_ref = "surgeon"

    if microlyte == "Yes":
        subject = f"Post-op incision innovation for {practice}"
        body = f"""Dr. {last},

{hook}, I wanted to introduce two innovations that are transforming post-operative care for spine surgeons.

First, ProPacks — standardized post-operative incision care kits designed for spine procedures:

- Better, more predictable patient outcomes with standardized wound care protocols
- Reduced physician time spent on post-op wound management
- Full compliance and audit protection with documented care pathways
- Increased per-case profitability through bundled incision care

Second, Microlyte SAM — an advanced antimicrobial wound matrix ideal for the longer incisions common in spine surgery:

- 99.99% microbial reduction sustained for 72+ hours (ionic and metallic silver)
- Fully synthetic and bioresorbable — no painful removal required
- Simple "peel and place" application over the full incision length
- HCPCS A2005 — reimbursable on Day 1 in your MAC jurisdiction

I'd love to set up a lunch & learn at {practice} in {city} to demonstrate both products — or a quick intro call if easier.

Would any day next week work for 15 minutes?

Best regards,
{sender['name']}
{sender['title']}
{sender['email']}"""
    else:
        subject = f"Post-op incision care for {practice}"
        body = f"""Dr. {last},

{hook}, I wanted to reach out about a solution helping {spec_ref}s like you improve post-op incision outcomes while saving time and increasing per-case profitability.

ProPacks are standardized post-operative incision care kits designed specifically for spine surgery patients. Surgeons using ProPacks report:

- Better, more predictable patient outcomes with standardized wound care protocols
- Reduced physician time spent on post-op wound management — critical with long spine incisions
- Full compliance and audit protection with documented care pathways
- Increased per-case profitability through bundled incision care

I'd love to set up a quick lunch & learn at {practice} in {city} to walk through how ProPacks could fit into your current post-op protocol — or a brief intro call works too.

Would any day next week work for a 15-minute conversation?

Best regards,
{sender['name']}
{sender['title']}
{sender['email']}"""

    return subject, body


def generate_outside_ortho_email(row, templates_config):
    """Generate outreach email for Outside Ortho & Spine leads."""
    sender = templates_config["sender"]
    last = row.get("Last Name", "")
    practice = row.get("Primary Site of Care") or "your practice"
    city = row.get("City") or ""
    specialty = str(row.get("Specialty") or "").lower()
    microlyte = str(row.get("Microlyte Eligible", "No"))

    vol = _safe_float(row.get("Procedure Vol"))

    # Detect specialty for personalization
    if "plastic" in specialty:
        spec_ref = "plastic surgeon"
        proc_ref = "reconstructive and cosmetic procedures"
    elif "vascular" in specialty:
        spec_ref = "vascular surgeon"
        proc_ref = "vascular procedures"
    elif "obstetric" in specialty or "gynecol" in specialty:
        spec_ref = "OB/GYN surgeon"
        proc_ref = "surgical procedures"
    elif "dermatol" in specialty or "mohs" in specialty:
        spec_ref = "dermatologic surgeon"
        proc_ref = "Mohs and excisional procedures"
    elif "colon" in specialty or "rectal" in specialty:
        spec_ref = "colorectal surgeon"
        proc_ref = "colorectal procedures"
    elif "oncol" in specialty:
        spec_ref = "surgical oncologist"
        proc_ref = "oncologic procedures"
    elif "transplant" in specialty:
        spec_ref = "transplant surgeon"
        proc_ref = "transplant procedures"
    else:
        spec_ref = "surgeon"
        proc_ref = "surgical procedures"

    if vol > 300:
        hook = f"With the volume of {proc_ref} you're managing at {practice}"
    elif vol >= 100:
        hook = f"Given your {proc_ref} practice at {practice}"
    else:
        hook = f"As a {spec_ref} performing {proc_ref} at {practice}"

    if microlyte == "Yes":
        subject = f"Post-op wound care innovation for {practice}"
        body = f"""Dr. {last},

{hook}, I wanted to introduce two innovations that are improving post-surgical wound outcomes across specialties.

First, ProPacks — standardized post-operative incision care kits:

- Better, more predictable patient outcomes with standardized wound care protocols
- Reduced physician time spent on post-op wound management
- Full compliance and audit protection with documented care pathways
- Increased per-case profitability through bundled incision care

Second, Microlyte SAM — an advanced antimicrobial wound matrix:

- 99.99% microbial reduction sustained for 72+ hours (ionic and metallic silver)
- Fully synthetic and bioresorbable — no painful removal required
- Simple "peel and place" application that integrates into any surgical workflow
- HCPCS A2005 — reimbursable on Day 1 in your MAC jurisdiction

I'd love to set up a lunch & learn at {practice} in {city} to demonstrate both products — or a quick intro call if easier.

Would any day next week work for 15 minutes?

Best regards,
{sender['name']}
{sender['title']}
{sender['email']}"""
    else:
        subject = f"Post-op incision care for {practice}"
        body = f"""Dr. {last},

{hook}, I wanted to reach out about a solution helping {spec_ref}s improve post-op incision outcomes while saving time and increasing per-case profitability.

ProPacks are standardized post-operative incision care kits that work across surgical specialties. Surgeons using ProPacks report:

- Better, more predictable patient outcomes with standardized wound care protocols
- Reduced physician time spent on post-op wound management
- Full compliance and audit protection with documented care pathways
- Increased per-case profitability through bundled incision care

I'd love to set up a quick lunch & learn at {practice} in {city} to walk through how ProPacks could fit into your current post-op protocol — or a brief intro call works too.

Would any day next week work for a 15-minute conversation?

Best regards,
{sender['name']}
{sender['title']}
{sender['email']}"""

    return subject, body


def write_enriched_tab(wb, sheet_name, df, vol_col, email_generator, templates_config):
    """Write an enriched tab with all pipeline data + tracking columns."""
    # Delete existing sheet if present
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # Compute priority
    vols = [_safe_float(r.get(vol_col)) for _, r in df.iterrows() if _safe_float(r.get(vol_col)) > 0]
    if vols:
        vols_sorted = sorted(vols)
        p25 = vols_sorted[len(vols_sorted) // 4]
        p75 = vols_sorted[3 * len(vols_sorted) // 4]
    else:
        p25, p75 = 0, 0

    df["_total_vol"] = df[vol_col].apply(_safe_float)
    df["Lead Priority"] = df.apply(lambda r: compute_priority(r, p25, p75), axis=1)
    df["Lead Status"] = "New"

    # Generate emails
    subjects = []
    drafts = []
    for _, row in df.iterrows():
        subj, body = email_generator(row, templates_config)
        subjects.append(subj)
        drafts.append(body)
    df["Subject Line"] = subjects
    df["Draft Email"] = drafts

    # Define output columns
    output_cols = [
        ("HCP NPI", 14), ("First Name", 13), ("Last Name", 15), ("Credential", 10),
        ("Specialty", 28), ("Verified Phone", 15), ("Phone Status", 16),
        ("Email", 32), ("Email Status", 22),
        ("Primary Site of Care", 32), ("Address", 25), ("City", 15),
        ("State", 8), ("Postal Code", 11), ("Practice Type", 16), ("Tier", 20),
        ("Lead Priority", 12), ("Lead Status", 16),
        ("MAC Jurisdiction", 16), ("Microlyte Eligible", 14),
        (vol_col, 14),
        ("Lg Collagen Vol", 14), ("Sm/Md Collagen Vol", 14), ("Collagen Powder Vol", 14),
        ("Lg Incision Likelihood", 16),
        ("Medical School", 28), ("HCP URL", 30),
        ("Subject Line", 35), ("Draft Email", 60),
        ("Call 1 Date", 12), ("Call 1 Outcome", 20), ("Call 1 Notes", 30),
        ("Call 2 Date", 12), ("Call 2 Outcome", 20), ("Call 2 Notes", 30),
        ("Call 3 Date", 12), ("Call 3 Outcome", 20), ("Call 3 Notes", 30),
        ("Email 1 Date", 12), ("Email 1 Outcome", 20), ("Email 1 Notes", 30),
        ("Email 2 Date", 12), ("Email 2 Outcome", 20), ("Email 2 Notes", 30),
        ("Next Action", 20), ("Next Action Date", 14),
        ("Decision Maker?", 14), ("Notes", 40),
    ]

    # Write headers
    for col_idx, (col_name, width) in enumerate(output_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # Write data sorted by volume desc
    df = df.copy()
    df["_sort"] = df[vol_col].apply(_safe_float)
    df = df.sort_values("_sort", ascending=False).drop(columns=["_sort", "_total_vol"])

    priority_col_idx = [i for i, (c, _) in enumerate(output_cols) if c == "Lead Priority"][0] + 1

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, (col_name, _) in enumerate(output_cols, start=1):
            val = row.get(col_name)
            if pd.isna(val) if isinstance(val, float) else val is None:
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

        # Priority color
        p = row.get("Lead Priority", "C")
        if p == "A":
            ws.cell(row=row_idx, column=priority_col_idx).fill = GREEN_FILL
        elif p == "B":
            ws.cell(row=row_idx, column=priority_col_idx).fill = YELLOW_FILL
        elif p == "C":
            ws.cell(row=row_idx, column=priority_col_idx).fill = RED_FILL

    # Auto-filter
    num_rows = len(df) + 1
    last_col = get_column_letter(len(output_cols))
    ws.auto_filter.ref = f"A1:{last_col}{num_rows}"

    # Dropdowns
    lead_status_col = get_column_letter([i for i, (c, _) in enumerate(output_cols) if c == "Lead Status"][0] + 1)
    dv_status = DataValidation(type="list", formula1='"New,Contacted,Meeting Scheduled,Meeting Completed,Proposal Sent,Won,Lost,Nurture"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_status)
    dv_status.add(f"{lead_status_col}2:{lead_status_col}{num_rows}")

    for call_outcome_col_name in ["Call 1 Outcome", "Call 2 Outcome", "Call 3 Outcome"]:
        idx = [i for i, (c, _) in enumerate(output_cols) if c == call_outcome_col_name]
        if idx:
            col_letter = get_column_letter(idx[0] + 1)
            dv = DataValidation(type="list", formula1='"Connected - Meeting Set,Connected - Call Back,Connected - Not Interested,Voicemail,No Answer,Wrong Number,Gatekeeper - Message Left,Gatekeeper - Blocked"', allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{num_rows}")

    for email_outcome_col_name in ["Email 1 Outcome", "Email 2 Outcome"]:
        idx = [i for i, (c, _) in enumerate(output_cols) if c == email_outcome_col_name]
        if idx:
            col_letter = get_column_letter(idx[0] + 1)
            dv = DataValidation(type="list", formula1='"Opened,Replied,Bounced,No Response,Unsubscribed"', allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{num_rows}")

    dm_idx = [i for i, (c, _) in enumerate(output_cols) if c == "Decision Maker?"]
    if dm_idx:
        col_letter = get_column_letter(dm_idx[0] + 1)
        dv = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{num_rows}")

    logger.info(f"Wrote {len(df)} leads to '{sheet_name}'")
    return ws


def main():
    input_file = "Master_Lead_List_Tracker (3).xlsx"
    output_file = "Master_Lead_List_Tracker (3).xlsx"

    # Load templates config
    with open("config/email_templates.json") as f:
        templates_config = json.load(f)

    # Extract data
    logger.info("Loading workbook...")
    wb_read = openpyxl.load_workbook(input_file, read_only=True)

    spine_df = extract_tab_to_df(wb_read, "Spine & Neuro")
    outside_df = extract_tab_to_df(wb_read, "Outside Ortho & Spine")
    wb_read.close()

    logger.info(f"Spine & Neuro: {len(spine_df)} leads")
    logger.info(f"Outside Ortho & Spine: {len(outside_df)} leads")

    # Enrich both
    spine_enriched = enrich_dataframe(spine_df, "Spine & Neuro")
    outside_enriched = enrich_dataframe(outside_df, "Outside Ortho & Spine")

    # Now open the workbook in write mode and replace the tabs
    logger.info("\nWriting enriched data back to workbook...")
    wb = openpyxl.load_workbook(input_file)

    write_enriched_tab(wb, "Spine & Neuro", spine_enriched, "Open Spine Vol",
                       generate_spine_neuro_email, templates_config)
    write_enriched_tab(wb, "Outside Ortho & Spine", outside_enriched, "Procedure Vol",
                       generate_outside_ortho_email, templates_config)

    # Reorder sheets to match original order
    desired_order = ["Dashboard", "Call Tracker - JR", "Email Tracker - JR",
                     "Spine & Neuro", "Outside Ortho & Spine", "Email Drafts - JR"]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(output_file)
    logger.info(f"\nSaved enriched workbook to {output_file}")

    # Also export CSVs for Google Sheets
    os.makedirs("data/output/csv_export", exist_ok=True)
    spine_enriched.to_csv("data/output/csv_export/Spine_Neuro.csv", index=False)
    outside_enriched.to_csv("data/output/csv_export/Outside_Ortho_Spine.csv", index=False)
    logger.info("CSVs exported to data/output/csv_export/")


if __name__ == "__main__":
    main()
